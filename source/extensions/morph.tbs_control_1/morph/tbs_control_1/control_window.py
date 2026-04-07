# SPDX-FileCopyrightText: Copyright (c) 2024 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: LicenseRef-NvidiaProprietary

"""
control_window.py — TBS 제어창 UI 및 이벤트 핸들러

【역할】
- build_control_window(ext): "TBS 제어창" 창. 최상단 화면 옵션(기본 메뉴/패널 숨기기),
  USD Load(load_window.build_load_ui_into_stack),
  USD 타임라인(수동/자동), 가상 시그널 샘플,
  XML 제너레이터(6종 시퀀스 콤보·입력 필드), 우선 표시 접두사, prim 목록.
- refresh_object_list(ext): 드롭다운/목록 갱신.
- on_play_usd_animation / on_play_generator_sample / on_refresh_prim_list 등 버튼·콤보 핸들러.

【수정 포인트】
- USD 재생 UI: build_control_window() 상단 ~ "USD 애니메이션 정지" 버튼 근처.
- XML 제너레이터 UI: "XML 제너레이터 생성기" Frame 블록.
  · 콤보 항목 추가/순서 변경: ext._xml_seq_combo = ui.ComboBox(0, ...) 인자 목록.
  · 하단 입력 전환: on_xml_seq_changed — FROM/TO 보일지(ext._xml_ab_inputs_frame), PORT만 보일지(ext._xml_port_inputs_frame).
    → xml_generator.FROM_TO_SEQS / PORT_ID_ONLY_SEQS 와 동일한 규칙 유지.
  · OK/역파싱: on_xml_ok_clicked, on_xml_run_clicked — 내부 seqs 리스트는 콤보 순서와 반드시 일치.
- 애니메이션 버튼(예: 이동/포물선/회전): 파일 하단 on_* 및 SAMPLE_GENERATOR_JSON.

【이벤트→XML→역파싱→애니메이션(JSON) 상태 기반 룰 매핑 유지보수 가이드】
- 목적: 같은 이벤트라도 from/to/포트 점유 상태가 다르면 다른 JSON을 실행할 수 있게 한다.
- 규칙 파일(우선순위):
  1) `config/event_animation_rules.json`  ← 권장(상태 기반)
  2) `config/event_animation_map.json`    ← 기본 fallback(이벤트 단순 매핑)
- 포트별 LOT prim 가시성: `config/port_lot_prim_paths.json` — `port_lot_visibility.apply_port_lot_prim_visibility` (시뮬 이벤트마다)
- rules 형식(요약):
  · 리스트 항목: {"name","priority","when","use"}
  · when:
    - sequence: 정식 시퀀스명(예: EAPEIS_PORT_MOVE_TRANSFERING)
    - from_port / to_port / port: 문자열 일치
    - ports_occupancy: {"EP2":"FULL","BP3":"EMPTY"} 같은 상태 조건
      (값은 FULL/EMPTY 또는 특정 LOT_ID)
  · use:
    - json: 실행할 시퀀스 JSON 경로
    - runner / description: 부가 메타
- 호출 흐름(현재 구현의 주 경로):
  1) simulation_engine._emit_event()에서 payload + ports_occupancy 전달
  2) on_sim_start_clicked()의 on_event 콜백이 post_sim_anim_event(...) → 큐 SimUiQueueKind.ANIM_EVENT
  3) _drain_sim_log_queue() → _sim_ui_sink_anim_event → handle_sim_event_for_animation(ext, payload)
  4) handle_sim_event_for_animation():
     - payload를 canonical sequence(EAPEIS_PORT_*)로 정규화
     - xml_generator.build_xml_string(...)로 XML 생성
     - xml_generator.parse_xml_string(...)으로 역파싱
     - 역파싱 결과(sequence_name/from/to/port)를 rules/map 입력 payload로 표준화
  5) _resolve_event_animation_entry(seq, payload):
     - 먼저 rules에서 조건 매칭(우선순위 높은 규칙 우선)
     - 없으면 event_animation_map fallback
  6) _execute_mapped_sequence_stub(...)에서 파일 존재/파싱 검증 후 SequenceRunner로 즉시 실행 시도
- 유지보수 체크포인트:
  · XML 생성/역파싱 규칙 수정: `xml_generator.py` (상수/빌더/파서)
  · 이벤트 별칭(canonical 변환) 수정: `SIM_SEQ_ALIAS`
  · rules 조건 필드 확장: `_resolve_rule_entry`, `_matches_occupancy_rule`
  · JSON 실행 연결/예외처리 수정: `_execute_mapped_sequence_stub`
  · 최종 분기 로그/표시 메시지 수정: `handle_sim_event_for_animation`
- 시퀀스 편집기 JSON 연결 방법:
  1) 시퀀스 편집기에서 JSON 저장
  2) 파일을 extension 내부 경로(예: data/sim_sequences/*.json)에 배치
  3) rules 또는 map의 use.json 경로에 등록
  4) 시뮬레이션 이벤트 발생 시 자동 매칭/검증 로그 확인
 - 표시모드(SimLogPanelMode): 콤보 인덱스와 `_drain_sim_log_queue`의 이력 스킵 여부가 연동된다.
  · "둘다": 진행현황 + 이력로그
  · "진행현황": 진행현황만
  · "이력로그": 스토리/시뮬 이력만
 - 시뮬 UI 큐 라우팅: `SimUiQueueKind` + `_dispatch_sim_ui_queue_item` + `_sim_ui_sink_*`.
   새 공정 텍스트 로그는 `post_sim_history_line(ext, line)`(시뮬 스레드)만 쓰면 이력 창으로 간다.
 - 시뮬레이션 종료 시 `_export_sim_logs_to_xlsx()`가 자동 호출되어
   `data/sim_logs/sim_logs_YYYYmmdd_HHMMSS.xlsx`에 2개 시트(진행현황/이력로그)를 저장한다.

【시뮬레이션 이벤트→애니메이션(JSON) 매핑 요약 (요구사항 반영)】
주의:
- 실제 실행 우선순위는 `EVENT_JSON_CASE_MAP` → rules(`config/event_animation_rules.json`) → map(`config/event_animation_map.json`) 순이다.
- 아래 목록은 코드 내 기본 테이블 `EVENT_JSON_CASE_MAP` 기준이며, payload의 키는 seq + (from_port_id/to_port_id 또는 port_id)이다.

1) 생성/투입(OHT 운반) — 애니는 ARRIVED에서만
- **이벤트(sequence_name)**: `EAPEIS_PORT_ARRIVED`
- **조건**: `from_port_id="OHT"`, `to_port_id="BP1|EP1|EP2|EP3"`
- **JSON**
  - OHT->BP1: `data/sim_sequences/arrived_bp1.json`
  - OHT->EP1: `data/sim_sequences/arrived_ep1.json`
  - OHT->EP2: `data/sim_sequences/arrived_ep2.json`
  - OHT->EP3: `data/sim_sequences/arrived_ep3.json`

2) BP1 → BP(버퍼) 이동 — 애니 실행
- **이벤트(sequence_name)**: `EAPEIS_PORT_MOVE_TRANSFERING`
- **조건**: `from_port_id="BP1"`, `to_port_id="BP2|BP3|BP4"`
- **JSON**
  - BP1->BP2: `data/sim_sequences/move_bp1_bp2.json`
  - BP1->BP3: `data/sim_sequences/move_bp1_bp3.json`
  - BP1->BP4: `data/sim_sequences/move_bp1_bp4.json`

3) BP → EP(공정포트) 이동 — 애니 실행
- **이벤트(sequence_name)**: `EISEAP_PORT_MOVE_REQ`
- **조건**: `from_port_id="BP2|BP3|BP4"`, `to_port_id="EP1|EP2|EP3"`
- **JSON**
  - BP2->EP1: `data/sim_sequences/move_bp2_ep1.json`
  - BP2->EP2: `data/sim_sequences/move_bp2_ep2.json`
  - BP2->EP3: `data/sim_sequences/move_bp2_ep3.json`
  - BP3->EP1: `data/sim_sequences/move_bp3_ep1.json`
  - BP3->EP2: `data/sim_sequences/move_bp3_ep2.json`
  - BP3->EP3: `data/sim_sequences/move_bp3_ep3.json`
  - BP4->EP1: `data/sim_sequences/move_bp4_ep1.json`
  - BP4->EP2: `data/sim_sequences/move_bp4_ep2.json`
  - BP4->EP3: `data/sim_sequences/move_bp4_ep3.json`

4) 회수 우선 실행 — 애니는 REMOVED에서만
- **이벤트(sequence_name)**: `EAPEIS_PORT_REMOVED`
- **조건**: `port_id="EP1|EP2|EP3"`
- **JSON**
  - EP1: `data/sim_sequences/removed_ep1.json`
  - EP2: `data/sim_sequences/removed_ep2.json`
  - EP3: `data/sim_sequences/removed_ep3.json`

5) 애니 없는 이벤트(상태/큐 의미만)
- `EAPEIS_PORT_READYTOLOAD` (생성/수신 준비)
- `EAPEIS_PORT_READYTOUNLOAD` (회수 요청 큐 적재)

【XML 시퀀스와 UI 필드】 (로직·상수는 xml_generator.py)
- FROM_PORT_ID + TO_PORT_ID: MOVE_TRANSFERING, MOVE, MOVE_REQ
- PORT_ID만: READYTOLOAD, ARRIVED, READYTOUNLOAD, REMOVED
새 종류 추가 시: xml_generator 수정 + 이 파일의 ComboBox·seqs 3곳 + 필요 시 IntField/모델 추가.

【주요 함수 색인(빠른 참조)】
- 경로·규칙 파일: _extension_root_dir(확장 루트), _event_animation_map_path / _event_animation_rules_path(JSON 경로),
  _load_event_animation_map·_load_event_animation_rules(mtime 캐시·로드)
- 규칙 매칭: _matches_occupancy_rule(ports_occupancy 조건), _resolve_rule_entry(rules.json 우선순위),
  _resolve_event_case_map_entry(EVENT_JSON_CASE_MAP), _resolve_event_animation_entry(통합: case→rules→map),
  _normalize_json_path(상대→절대)
- 실행: _execute_mapped_sequence_stub(매핑된 JSON 검증·SequenceRunner 실행), _estimate_step_duration_sec_for_log,
  _estimate_sequence_total_duration_sec_for_log, _estimate_anim_duration_for_gate_payload(게이트 대기 시간 추정)
- 이벤트 처리: handle_sim_event_for_animation(시뮬 payload→XML→역파싱→룰→JSON 실행), _on_sim_event(로그용 래퍼)
- 시뮬 UI 큐(스레드→UI): post_sim_history_line, post_sim_anim_event, post_sim_progress_update,
  _enqueue_sim_log·_enqueue_anim_event·_enqueue_control_action·_enqueue_gate_request·_enqueue_sim_progress,
  _drain_sim_log_queue(메인 스레드에서 소비), _dispatch_sim_ui_queue_item, _coerce_sim_ui_queue_kind,
  _sim_ui_sink_progress·_sim_ui_sink_anim_event·_sim_ui_sink_history_line·_sim_ui_sink_action·_sim_ui_sink_gate
- 게이트: _show_sim_gate_dialog, _close_sim_gate_dialog, on_sim_start_clicked 내부 on_gate 연동
- 진행·로그 UI: _append_sim_log, _format_history_line·_with_history_color_icon, _append_anim_history_log(노옵),
  _render_pending_dots,
  _update_sim_progress, _is_progress_only_mode, on_copy_sim_progress
- 포트 패널: _port_cell_text, _compact_cell_value, _sync_ep3_port_cell_visibility, _set_port_box_style, _update_port_occupancy_panel
- 시뮬 제어: on_sim_start_clicked·on_sim_stop_clicked·on_sim_reset_clicked, _detach_sim_update,
  on_sim_log_view_changed, on_sim_ep_count_changed, _export_sim_logs_to_xlsx
- XML UI: on_xml_seq_changed, on_xml_ok_clicked, on_xml_run_clicked
- 포트 문자열: _parse_port_num, _port_kind, _normalize_port_text_from_xml
- Prim 목록: on_refresh_prim_list, refresh_object_list, build_object_panel, on_button_0/1/2
- 가상 시그널: receive_signal_data, run_generator_from_parsed
- 창: build_control_window(전체 UI 조립)

사용처: extension.py on_startup → build_control_window(self)
  · 재호출 시 기존 TBS 제어창은 destroy 후 재생성(확장 리로드 등으로 위젯 이중 생성 방지).
"""

from enum import Enum
from typing import Any, Dict, List, Optional, Tuple
import random
import threading
import time
import queue
import json
import re
from datetime import datetime
from pathlib import Path

import omni.kit.app as app
import omni.ui as ui
from pxr import Gf

from . import usd_animation_control
from . import xml_generator
from .curve_animation import make_parabolic_path, run_prim_curve_animation, stop_prim_curve_animation
from .kit_chrome_visibility import KIT_CHROME_HIDE_DEFAULT_ON_LAUNCH, apply_kit_chrome_hidden
from .load_window import build_load_ui_into_stack
from .port_lot_visibility import apply_port_lot_prim_visibility, clear_port_lot_authoring_cache
from .prim_info import get_prim_display_name, safe_str
from .prim_utils import (
    collect_prim_paths_safe,
    find_all_prim_paths_by_name,
    get_prim_local_translate,
    get_stage,
    set_prim_translate_only,
)
from .rotate_animation import run_prim_rotate_animation, stop_prim_rotate_animation
from .selection_overlay import show_prim_info_in_viewport
from .signal_parser import parse_signal
from .sequence_engine import SequenceRunner
from .simulation_engine import (
    Lot,
    SimulationInitConfig,
    SimulationLogConfig,
    SimulationTimingConfig,
    TBSSimulationEngine,
)
from .translate_animation import run_prim_translate_animation, stop_prim_translate_animation

MAX_PRIMS_DISPLAY = 80
DEFAULT_PRIORITY_NAME_PREFIX = "Mesh_"
CHECKBOX_WHITE_STYLE = {
    "color": 0xFF000000,
    "background_color": 0xFFEEEEEE,
}


class SimUiQueueKind(str, Enum):
    """
    `_sim_log_queue` 튜플 (kind, payload) 의 kind.
    payload가 도달하는 UI 영역은 아래 sink와 1:1에 가깝게 대응한다.
    """

    PROGRESS = "progress"  # → 진행현황 패널 (_update_sim_progress)
    HISTORY_LINE = "log"  # → 이력 로그 패널 (_append_sim_log). 값 "log"는 기존 큐 호환 유지.
    ANIM_EVENT = "anim_event"  # → 포트 상태 + 시퀀스 실행 (handle_sim_event_for_animation)
    ACTION = "action"  # → 제어 액션 (예: xlsx보내기)
    GATE = "gate"  # → 공정 확인 창


class SimUiControlAction(str, Enum):
    """SimUiQueueKind.ACTION 의 payload 로 허용되는 값."""

    EXPORT_XLSX = "export_xlsx"


class SimLogPanelMode(int, Enum):
    """표시모드 콤보 인덱스 (`on_sim_log_view_changed` 와 동일)."""

    ALL = 0
    PROGRESS_ONLY = 1
    HISTORY_ONLY = 2


SIM_SEQ_ALIAS = {
    "READYTOLOAD": xml_generator.SEQ_READYTOLOAD,
    "ARRIVED": xml_generator.SEQ_ARRIVED,
    "MOVE_TRANSFERING": xml_generator.SEQ_MOVE_TRANSFERING,
    "MOVE": xml_generator.SEQ_MOVE,
    "MOVE_REQ": xml_generator.SEQ_MOVE_REQ,
    "READYTOUNLOAD": xml_generator.SEQ_READYTOUNLOAD,
    "REMOVED": xml_generator.SEQ_REMOVED,
}
# 이벤트별/포트별 JSON 매핑(최상단 일원화)
# - 운영 중 수정은 이 테이블을 우선 수정한다.
# - key 규칙:
#   * READYTOLOAD/READYTOUNLOAD: 애니 없음(매핑 비워둠)
#   * ARRIVED: OHT 이동 애니만 실행 → key="FROM->TO" (예: OHT->EP1, OHT->BP1)
#   * EAPEIS_PORT_MOVE_TRANSFERING: BP1->BPx 이동 애니만 실행 → key="FROM->TO"
#   * EISEAP_PORT_MOVE_REQ: BPx->EPy 이동 애니만 실행 → key="FROM->TO"
#   * REMOVED: 회수 우선순위가 되었을 때 회수 애니 실행 → key="PORT" (EP1/2/3)
EVENT_JSON_CASE_MAP: Dict[str, Dict[str, str]] = {
    # 요구사항: OHT 이동은 ARRIVED에서만(=MOVE 불필요)
    xml_generator.SEQ_MOVE_TRANSFERING: {
        "BP1->BP2": "data/sim_sequences/move_bp1_bp2.json",  # BP1 LOT를 BP2 버퍼로 이송
        "BP1->BP3": "data/sim_sequences/move_bp1_bp3.json",  # BP1 LOT를 BP3 버퍼로 이송
        "BP1->BP4": "data/sim_sequences/move_bp1_bp4.json",  # BP1 LOT를 BP4 버퍼로 이송
    },
    xml_generator.SEQ_ARRIVED: {
        # 요구사항: ARRIVED는 "OHT가 옮기는 모든 애니" 트리거. FROM->TO 로 매핑한다.
        "OHT->BP1": "data/sim_sequences/arrived_bp1.json",
        "OHT->EP1": "data/sim_sequences/arrived_ep1.json",
        "OHT->EP2": "data/sim_sequences/arrived_ep2.json",
        "OHT->EP3": "data/sim_sequences/arrived_ep3.json",
    },
    # 요구사항: READYTOLOAD / READYTOUNLOAD 는 애니 실행 안함(빈 dict 유지)
    xml_generator.SEQ_READYTOLOAD: {},
    xml_generator.SEQ_READYTOUNLOAD: {},
    # 요구사항: BP->EP 이동 애니는 EISEAP_PORT_MOVE_REQ 에서만 실행
    xml_generator.SEQ_MOVE_REQ: {
        "BP2->EP1": "data/sim_sequences/move_bp2_ep1.json",
        "BP2->EP2": "data/sim_sequences/move_bp2_ep2.json",
        "BP2->EP3": "data/sim_sequences/move_bp2_ep3.json",
        "BP3->EP1": "data/sim_sequences/move_bp3_ep1.json",
        "BP3->EP2": "data/sim_sequences/move_bp3_ep2.json",
        "BP3->EP3": "data/sim_sequences/move_bp3_ep3.json",
        "BP4->EP1": "data/sim_sequences/move_bp4_ep1.json",
        "BP4->EP2": "data/sim_sequences/move_bp4_ep2.json",
        "BP4->EP3": "data/sim_sequences/move_bp4_ep3.json",
    },
    xml_generator.SEQ_REMOVED: {
        "EP1": "data/sim_sequences/removed_ep1.json",  # EP1에서 LOT/Foup 회수 완료 연출
        "EP2": "data/sim_sequences/removed_ep2.json",  # EP2에서 LOT/Foup 회수 완료 연출
        "EP3": "data/sim_sequences/removed_ep3.json",  # EP3에서 LOT/Foup 회수 완료 연출
    },
}
SAMPLE_GENERATOR_JSON = """{
  "objects": ["Mesh_308", "Mesh_561", "WalkwayEndA_01"],
  "animation": {
    "segments": [
      {"duration": 1.0, "delta": [100, 0, 0]},
      {"duration": 1.0, "delta": [0, 100, 0]},
      {"duration": 2.0, "delta": [-100, -100, 0]}
    ]
  }
}"""

_EVENT_ANIM_MAP_CACHE: Optional[Dict[str, Any]] = None
_EVENT_ANIM_MAP_MTIME: Optional[float] = None
_EVENT_ANIM_RULES_CACHE: Optional[List[Dict[str, Any]]] = None
_EVENT_ANIM_RULES_MTIME: Optional[float] = None


def _extension_root_dir() -> Path:
    """이 파일 기준 확장 루트(.../morph.tbs_control_1). config·data 경로 계산에 사용."""
    # .../source/extensions/morph.tbs_control_1
    return Path(__file__).resolve().parents[2]


def _event_animation_map_path() -> Path:
    """이벤트 seq → JSON 단순 매핑 파일 경로."""
    return _extension_root_dir() / "config" / "event_animation_map.json"


def _event_animation_rules_path() -> Path:
    """상태 기반 애니메이션 규칙(우선순위 리스트) 파일 경로."""
    return _extension_root_dir() / "config" / "event_animation_rules.json"


def _load_event_animation_map() -> Dict[str, Any]:
    """event_animation_map.json을 읽어 dict로 반환. mtime이 같으면 캐시 재사용."""
    global _EVENT_ANIM_MAP_CACHE, _EVENT_ANIM_MAP_MTIME
    p = _event_animation_map_path()
    if not p.exists():
        _EVENT_ANIM_MAP_CACHE = {}
        _EVENT_ANIM_MAP_MTIME = None
        return {}
    try:
        mtime = p.stat().st_mtime
        if _EVENT_ANIM_MAP_CACHE is not None and _EVENT_ANIM_MAP_MTIME == mtime:
            return _EVENT_ANIM_MAP_CACHE
        data = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            data = {}
        _EVENT_ANIM_MAP_CACHE = data
        _EVENT_ANIM_MAP_MTIME = mtime
        return data
    except Exception as e:
        print(f"[ANIM MAP] 매핑 파일 로드 실패: {p} err={e}", flush=True)
        _EVENT_ANIM_MAP_CACHE = {}
        _EVENT_ANIM_MAP_MTIME = None
        return {}


def _load_event_animation_rules() -> List[Dict[str, Any]]:
    """event_animation_rules.json을 읽어 규칙 리스트로 반환(priority 오름차순 정렬, mtime 캐시)."""
    global _EVENT_ANIM_RULES_CACHE, _EVENT_ANIM_RULES_MTIME
    p = _event_animation_rules_path()
    if not p.exists():
        _EVENT_ANIM_RULES_CACHE = []
        _EVENT_ANIM_RULES_MTIME = None
        return []
    try:
        mtime = p.stat().st_mtime
        if _EVENT_ANIM_RULES_CACHE is not None and _EVENT_ANIM_RULES_MTIME == mtime:
            return _EVENT_ANIM_RULES_CACHE
        data = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            data = []
        norm: List[Dict[str, Any]] = [x for x in data if isinstance(x, dict)]
        norm.sort(key=lambda r: int(r.get("priority", 1000)))
        _EVENT_ANIM_RULES_CACHE = norm
        _EVENT_ANIM_RULES_MTIME = mtime
        return norm
    except Exception as e:
        print(f"[ANIM RULES] 규칙 파일 로드 실패: {p} err={e}", flush=True)
        _EVENT_ANIM_RULES_CACHE = []
        _EVENT_ANIM_RULES_MTIME = None
        return []


def _matches_occupancy_rule(rule_occ: Dict[str, Any], occ: Dict[str, Any]) -> bool:
    """규칙의 ports_occupancy(포트→FULL/EMPTY/LOT_ID)가 현재 occ 스냅샷과 일치하는지."""
    for port, expected in (rule_occ or {}).items():
        p = str(port).strip().upper()
        got = str((occ or {}).get(p, "") or "")
        exp = str(expected or "").strip()
        if exp.upper() == "FULL":
            if not got:
                return False
        elif exp.upper() == "EMPTY":
            if got:
                return False
        else:
            if got != exp:
                return False
    return True


def _resolve_rule_entry(seq: str, payload: Dict[str, str]) -> Tuple[Optional[str], Optional[Dict[str, Any]], Optional[str]]:
    rules = _load_event_animation_rules()
    if not rules:
        return (None, None, None)
    p_from = str(payload.get("from_port_id", "") or "")
    p_to = str(payload.get("to_port_id", "") or "")
    p_port = str(payload.get("port_id", "") or "")
    occ = payload.get("ports_occupancy", {}) if isinstance(payload.get("ports_occupancy", {}), dict) else {}

    for r in rules:
        when = r.get("when", {}) if isinstance(r.get("when", {}), dict) else {}
        if str(when.get("sequence", "") or "").strip() not in ("", seq):
            continue
        if str(when.get("from_port", "") or "").strip() not in ("", p_from):
            continue
        if str(when.get("to_port", "") or "").strip() not in ("", p_to):
            continue
        if str(when.get("port", "") or "").strip() not in ("", p_port):
            continue
        rule_occ = when.get("ports_occupancy", {})
        if isinstance(rule_occ, dict) and not _matches_occupancy_rule(rule_occ, occ):
            continue

        use = r.get("use", {}) if isinstance(r.get("use", {}), dict) else {}
        j = use.get("json")
        if isinstance(j, str) and j.strip():
            return (j.strip(), use, str(r.get("name", "")))
    return (None, None, None)


def _resolve_event_case_map_entry(seq: str, payload: Dict[str, str]) -> Tuple[Optional[str], Optional[Dict[str, Any]], Optional[str]]:
    table = EVENT_JSON_CASE_MAP.get(seq, {})
    if not isinstance(table, dict) or not table:
        return (None, None, None)
    p_from = str(payload.get("from_port_id", "") or "").strip().upper()
    p_to = str(payload.get("to_port_id", "") or "").strip().upper()
    p_port = str(payload.get("port_id", "") or "").strip().upper()
    key = f"{p_from}->{p_to}" if p_from and p_to else p_port
    if not key:
        return (None, None, None)
    j = table.get(key)
    if not isinstance(j, str) or not j.strip():
        return (None, None, None)
    meta = {
        "runner": "sequence_editor",
        "description": f"top-case-map:{key}",
    }
    return (j.strip(), meta, f"top_case_map:{seq}:{key}")


def _resolve_event_animation_entry(seq: str, payload: Optional[Dict[str, str]] = None) -> Tuple[Optional[str], Optional[Dict[str, Any]], Optional[str], str]:
    """
    반환:
    - json_path_str: 실제 JSON 경로 문자열(없으면 None)
    - meta: runner/description 등 부가정보
    """
    # 0) 파일 최상단 케이스 매핑(운영 우선)
    j_case, meta_case, case_name = _resolve_event_case_map_entry(seq, payload or {})
    if j_case:
        return (j_case, meta_case, case_name, "top_case_map")

    # 1) 상태 기반 rules 우선
    j_rule, meta_rule, rule_name = _resolve_rule_entry(seq, payload or {})
    if j_rule:
        return (j_rule, meta_rule, rule_name or "(unnamed-rule)", "rules")

    # 2) 기존 단순 map fallback
    m = _load_event_animation_map()
    if not m:
        return (None, None, None, "")

    # 키 우선순위:
    # 1) 정식 seq (EAPEIS_PORT_...)
    # 2) 별칭(READYTOLOAD 등)
    # 3) raw 값
    raw_alias = None
    for alias, canonical in SIM_SEQ_ALIAS.items():
        if canonical == seq:
            raw_alias = alias
            break

    cand = [seq]
    if raw_alias:
        cand.append(raw_alias)

    for key in cand:
        if key not in m:
            continue
        v = m.get(key)
        if isinstance(v, str):
            return (v, {"runner": "sequence_editor", "description": ""}, None, "map")
        if isinstance(v, dict):
            j = v.get("json")
            if isinstance(j, str) and j.strip():
                return (j.strip(), v, None, "map")
    return (None, None, None, "")


def _normalize_json_path(path_text: str) -> Path:
    """시퀀스 JSON 상대 경로를 확장 루트 기준 절대 Path로."""
    p = Path(path_text)
    if p.is_absolute():
        return p
    return (_extension_root_dir() / p).resolve()


def _estimate_step_duration_sec_for_log(step: Dict[str, Any]) -> Optional[float]:
    """
    애니메이션 실행이력용 "예상 길이" 계산(보수적).
    - MOVE/ROTATE: duration_max가 있으면 max, 없으면 duration.
    - DELAY: duration
    - USD_TIMELINE: 수동이면 프레임 범위로 추정, AUTO는 환경 의존이라 None.
    """
    try:
        t = str((step or {}).get("type") or "").upper()
    except Exception:
        return None
    try:
        if t in ("MOVE", "ROTATE"):
            if "duration_max" in (step or {}):
                return max(0.0, float((step or {}).get("duration_max", (step or {}).get("duration", 0.0))))
            return max(0.0, float((step or {}).get("duration", 0.0)))
        if t == "DELAY":
            return max(0.0, float((step or {}).get("duration", 0.0)))
        if t == "USD_TIMELINE":
            mode = str((step or {}).get("mode", "MANUAL")).upper()
            if mode == "AUTO":
                return None
            start = int((step or {}).get("start_frame", 0))
            end = int((step or {}).get("end_frame", 0))
            if end <= start:
                return 0.0
            # 대략치: 60fps 가정(Kit 환경별로 다를 수 있어 "예상치"로만 사용)
            return max(0.0, float(end - start) / 60.0)
    except Exception:
        return None
    return None


def _estimate_sequence_total_duration_sec_for_log(steps: List[Dict[str, Any]]) -> Optional[float]:
    """
    SequenceRunner의 그룹/지연 규칙을 단순화해서 "예상 총 길이"를 계산한다.
    - 병렬 그룹: 리더 시작 시각 기준으로 (offset + duration)의 최대값을 그룹 종료로 본다.
    - 다음 그룹 시작: engine과 동일하게 anchor_end + next.step_delay_ms 를 사용하되,
      그룹 시작(t0)보다 앞당기지는 않는다.
    """
    if not steps:
        return 0.0
    # 첫 스텝의 step_delay_ms는 시퀀스 시작 전 지연으로 해석
    try:
        t_cursor = max(0.0, int((steps[0] or {}).get("step_delay_ms", 0)) / 1000.0)
    except Exception:
        t_cursor = 0.0
    last_finish = t_cursor

    i = 0
    while i < len(steps):
        try:
            g_end = _group_end_index(steps, i)
        except Exception:
            g_end = i
        t0 = t_cursor

        # 그룹 내 예상 종료(병렬 최대)
        group_finish = t0
        for j in range(i, g_end + 1):
            st = steps[j] if isinstance(steps[j], dict) else {}
            off = 0.0
            if j != i:
                try:
                    off = max(0.0, int((st or {}).get("step_delay_ms", 0)) / 1000.0)
                except Exception:
                    off = 0.0
            dur = _estimate_step_duration_sec_for_log(st)
            if dur is None:
                # 알 수 없는 타입/auto 타임라인이 섞이면 전체 추정도 None 처리
                return None
            group_finish = max(group_finish, t0 + off + float(dur))
        last_finish = max(last_finish, group_finish)

        next_idx = g_end + 1
        if next_idx >= len(steps):
            break

        # anchor 종료 시각(앵커 스텝은 그룹 마지막)
        anchor_step = steps[g_end] if isinstance(steps[g_end], dict) else {}
        anchor_off = 0.0
        if g_end > i:
            try:
                anchor_off = max(0.0, int((anchor_step or {}).get("step_delay_ms", 0)) / 1000.0)
            except Exception:
                anchor_off = 0.0
        anchor_dur = _estimate_step_duration_sec_for_log(anchor_step)
        if anchor_dur is None:
            return None
        anchor_end = t0 + anchor_off + float(anchor_dur)

        try:
            delay_next = int((steps[next_idx] or {}).get("step_delay_ms", 0)) / 1000.0
        except Exception:
            delay_next = 0.0
        t_cursor = max(t0, anchor_end + float(delay_next))
        i = next_idx

    return max(0.0, float(last_finish))


def _execute_mapped_sequence_stub(
    ext: Any,
    seq: str,
    payload: Dict[str, str],
    json_path_text: str,
    meta: Optional[Dict[str, Any]],
    rule_name: Optional[str],
    verbose: bool,
) -> None:
    """
    rules/map이 가리키는 JSON을 검증한 뒤 SequenceRunner.run()으로 실제 재생한다.
    시뮬 tick은 (필요 시) translate/rotate/curve 애니 재생 중에만 _sim_tick_pause_event로 잠시 멈춘다.
    JSON 시퀀스(SequenceRunner) 재생 중에는 tick을 멈추지 않아 공정(sim time)과 동시에 진행된다.
    """
    p = _normalize_json_path(json_path_text)
    runner = str((meta or {}).get("runner", "sequence_editor"))
    desc = str((meta or {}).get("description", ""))
    from_port = str(payload.get("from_port_id", "")).strip()
    to_port = str(payload.get("to_port_id", "")).strip()
    port = str(payload.get("port_id", "")).strip()
    if from_port and to_port:
        route = f"{from_port}->{to_port}"
    elif to_port:
        route = f"to={to_port}"
    elif from_port:
        route = f"from={from_port}"
    elif port:
        route = f"port={port}"
    else:
        route = "port=미상"
    base_desc = desc if desc else "동작설명 없음"
    lot_id = str(payload.get("lot_id", "")).strip() or "-"
    action_text = f"{base_desc} ({route} | lot={lot_id})"
    sim_time = str(payload.get("sim_time", "")).strip()
    # 스토리-JSON 요약(애니 실행이력 창) 기록: "실행/스킵/실패" 모두 남겨야 누적이 끊기지 않는다.
    def _push_story_json(status: str) -> None:
        """(애니 실행이력 UI 제거) 스토리-JSON 요약 누적은 더 이상 사용하지 않는다."""
        return

    if not p.exists():
        _push_story_json("SKIP (MISSING)")
        _append_anim_history_log(
            ext,
            f"[ANIM] 파일없음 | event={seq} | action={action_text} | need={p.name} | path={p}",
        )
        if verbose:
            print(
                f"[ANIM MAP] 이벤트={seq} -> JSON 파일 없음: {p} "
                f"(runner={runner}, rule={rule_name or '-'}, desc={desc})",
                flush=True,
            )
        return
    try:
        # 파일 유효성 확인(실행 전 파싱 검증)
        parsed = json.loads(p.read_text(encoding="utf-8"))
        if not isinstance(parsed, list):
            raise ValueError("시퀀스 JSON 루트는 list여야 합니다.")
    except Exception as e:
        _push_story_json("SKIP (PARSE_FAIL)")
        _append_anim_history_log(ext, f"[ANIM] JSON 파싱실패 | event={seq} | action={action_text} | file={p.name} | err={e}")
        if verbose:
            print(f"[ANIM MAP] JSON 파싱 실패: {p} err={e}", flush=True)
        return

    # 시뮬레이션 중에는 "빈 JSON([])"을 실행(run)하면 runner 초기화/복원 경로를 타면서
    # 포즈/숨김이 튀는 원인이 된다. 빈 시퀀스는 실행하지 않고 스킵한다.
    if not parsed:
        _push_story_json("SKIP (EMPTY)")
        _append_anim_history_log(
            ext,
            f"[ANIM] 스킵(EMPTY JSON) | t={sim_time or '-'} | event={seq} | action={action_text} | file={p.name}",
        )
        if verbose:
            print(f"[ANIM MAP] 빈 JSON([]) 스킵: {p} event={seq}", flush=True)
        return

    # 예상 총 길이(초): 엑셀/로그에 같이 남길 수 있게 추정
    est_total = _estimate_sequence_total_duration_sec_for_log(parsed)
    est_text = f"{est_total:.2f}s" if isinstance(est_total, (float, int)) else "미확인"

    step_types: List[str] = []
    for step in parsed:
        if isinstance(step, dict):
            t = str(step.get("type", "")).strip().upper()
            if t:
                step_types.append(t)
    if step_types:
        preview = ",".join(step_types[:4]) + ("..." if len(step_types) > 4 else "")
    else:
        preview = "EMPTY"

    _append_anim_history_log(
        ext,
        f"[ANIM] 실행준비완료 | t={sim_time or '-'} | event={seq} | est={est_text} | action={action_text} | file={p.name} | steps={len(parsed)}({preview}) | runner={runner} | rule={rule_name or '-'}",
    )
    _push_story_json("PLAN (READY)")
    # 실제 실행 연결: 동시에 여러 run() 호출 시 기존 애니가 끊기는 문제를 막기 위해
    # "시뮬레이션 애니 실행 큐"로 직렬화한다.
    try:
        if not isinstance(getattr(ext, "_sim_anim_pending", None), list):
            ext._sim_anim_pending = []

        def _start_job(job: Dict[str, Any]) -> None:
            pause_evt = getattr(ext, "_sim_tick_pause_event", None)
            started_wall = time.monotonic()
            active = dict(job)
            active["_started_wall"] = started_wall
            ext._sim_anim_active = active
            try:
                if isinstance(job.get("est_total"), (float, int)) and float(job.get("est_total")) > 0.0:
                    ext._sim_tick_pause_until_wall = float(started_wall) + float(job.get("est_total"))
                else:
                    ext._sim_tick_pause_until_wall = None
            except Exception:
                ext._sim_tick_pause_until_wall = None

            def _on_done():
                # 포트상태 점(●) 감소 시점
                # - ARRIVED(OHT->*) 애니가 "포트 도착"을 의미하므로, 완료 후 생성 토큰 1개 소모
                # - REMOVED 애니가 "회수 진행"이므로, 완료 후 회수 토큰 1개 소모
                # (요청으로 제거) 포트상태 좌/우 점 표시 기능 비활성화
                # 정책 변경:
                # - 애니메이션 완료 후에는 "완료된 자세 그대로" 유지한다.
                # - 다음 애니메이션 시작 시점에만 시퀀서 stop() 경로(=baseline 복원/초기화)가 동작하도록,
                #   완료 직후 baseline을 현재 자세로 덮어쓰지 않는다.
                pending = getattr(ext, "_sim_anim_pending", [])
                if isinstance(pending, list) and pending:
                    # 우선순위 큐: _priority 낮은 job 먼저
                    try:
                        pending.sort(key=lambda j: int((j or {}).get("_priority", 10)) if isinstance(j, dict) else 10)
                    except Exception:
                        pass
                    nxt = pending.pop(0)
                    _start_job(nxt)
                    return
                if pause_evt is not None:
                    try:
                        pause_evt.clear()
                    except Exception:
                        pass
                try:
                    ext._sim_tick_pause_until_wall = None
                except Exception:
                    pass
                try:
                    _refresh_sim_progress_from_last(ext)
                except Exception:
                    pass

            try:
                ext._sim_runner.on_sequence_completed = _on_done  # type: ignore[attr-defined]
            except Exception:
                pass

            sp = 1.0
            try:
                m = getattr(ext, "_sim_speed_model", None)
                if m is not None:
                    sp = max(0.1, float(m.get_value_as_float()))
            except Exception:
                sp = 1.0
            if sp > 1.0 and pause_evt is not None:
                print("호출됨!!!!")
                try:
                    pause_evt.set()
                except Exception:
                    pass
            # JSON 시퀀스 재생 중에도 sim tick이 돌아가야 _wait_with_progress(공정)와 애니가 동시에 진행된다.
            # 배속>1일 때만 pause_evt.set()으로 tick을 잠시 맞춤(1배속에서는 set 하지 않음).
            ext._sim_runner.run(job.get("parsed", []))
            try:
                _refresh_sim_progress_from_last(ext)
            except Exception:
                pass

        job = {
            "t": sim_time,
            "event": seq,
            "file": p.name,
            "path": str(p),
            "action": action_text,
            "est": est_text,
            "est_total": float(est_total) if isinstance(est_total, (float, int)) else None,
            "runner": runner,
            "rule": rule_name or "-",
            "lot_id": lot_id,
            "from_port_id": from_port,
            "to_port_id": to_port,
            "port_id": port,
            "parsed": parsed,
        }
        # 우선순위: 생성(OHT->EP 직접투입 등) / 회수(REMOVED) 는 현재 애니가 끝나자마자 즉시 실행되어야 한다.
        # - 선점(interrupt)은 하지 않고, pending 큐의 "앞"에 삽입한다.
        try:
            is_pickup = str(seq).strip().upper() == str(xml_generator.SEQ_REMOVED).strip().upper()
        except Exception:
            is_pickup = False
        try:
            is_spawn = (
                str(seq).strip().upper() == str(xml_generator.SEQ_ARRIVED).strip().upper()
                and str(from_port).strip().upper() == "OHT"
                and str(to_port).strip().upper().startswith("EP")
            )
        except Exception:
            is_spawn = False
        job["_priority"] = 0 if (is_spawn or is_pickup) else 10
        try:
            runner_busy = bool(
                getattr(ext, "_sim_runner", None) is not None
                and getattr(ext._sim_runner, "is_running", lambda: False)()
            )
        except Exception:
            runner_busy = False
        if runner_busy:
            try:
                pending = getattr(ext, "_sim_anim_pending", None)
                if not isinstance(pending, list):
                    pending = []
                    ext._sim_anim_pending = pending
                if int(job.get("_priority", 10)) <= 0:
                    pending.insert(0, job)
                else:
                    pending.append(job)
            except Exception:
                ext._sim_anim_pending.append(job)
            _append_anim_history_log(
                ext,
                f"[ANIM] 대기큐적재 | event={seq} | est={est_text} | action={action_text} | file={p.name} | queued={len(ext._sim_anim_pending)}",
            )
            try:
                _refresh_sim_progress_from_last(ext)
            except Exception:
                pass
            return
        _start_job(job)
    except Exception as e:
        _append_anim_history_log(ext, f"[ANIM] 실행실패 | event={seq} | action={action_text} | file={p.name} | err={e}")
        pause_evt = getattr(ext, "_sim_tick_pause_event", None)
        if pause_evt is not None:
            try:
                pause_evt.clear()
            except Exception:
                pass
        try:
            ext._sim_tick_pause_until_wall = None
        except Exception:
            pass

    if verbose:
        print(
            f"[ANIM MAP] 이벤트={seq} -> JSON 준비완료: {p} "
            f"(runner={runner}, rule={rule_name or '-'}, lot={payload.get('lot_id','')}, port={payload.get('port_id','')}, "
            f"from={payload.get('from_port_id','')}, to={payload.get('to_port_id','')})",
            flush=True,
        )


def _estimate_anim_duration_for_gate_payload(ext: Any, payload: Dict[str, str]) -> float:
    """
    simulation_engine의 on_gate에서 호출되는 "애니메이션 예상 길이" 계산기.
    - 게이트 시점에 XML 생성/역파싱 → rules/map 매핑 → JSON 파싱 → 총 duration 추정
    - 실패하면 0.0 반환(=애니 대기 없음)
    """
    try:
        seq_raw = str(payload.get("seq", "") or "").strip()
        if not seq_raw:
            return 0.0
        seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
        fr = str(payload.get("from_port_id", "") or "")
        to = str(payload.get("to_port_id", "") or "")
        port = str(payload.get("port_id", "") or "")

        # 1차: 원본 payload 기준(최소한 map fallback은 항상 시도)
        mapping_payload = dict(payload or {})
        mapping_payload["seq"] = seq

        # 2차: XML 표준화가 가능하면 덮어쓴다(우선 적용)
        try:
            if seq in xml_generator.FROM_TO_SEQS:
                xml_text = xml_generator.build_xml_string(
                    seq,
                    from_port_id=_parse_port_num(fr, 1),
                    to_port_id=_parse_port_num(to, 1),
                )
            elif seq in xml_generator.PORT_ID_ONLY_SEQS:
                xml_text = xml_generator.build_xml_string(seq, port_id=_parse_port_num(port, 1))
            else:
                xml_text = ""
            if xml_text:
                parsed = xml_generator.parse_xml_string(xml_text) or {}
                seq_for_mapping = str(parsed.get("sequence_name", "") or "").strip().upper() or seq
                mapping_payload["seq"] = seq_for_mapping
                mapping_payload["from_port_id"] = _normalize_port_text_from_xml(str(parsed.get("from_port_id", "") or ""), fr)
                mapping_payload["to_port_id"] = _normalize_port_text_from_xml(str(parsed.get("to_port_id", "") or ""), to)
                mapping_payload["port_id"] = _normalize_port_text_from_xml(str(parsed.get("port_id", "") or ""), port)
                seq = seq_for_mapping
        except Exception:
            # XML 표준화 실패해도 원본 payload로 rules/map 추정을 계속 시도한다.
            pass

        # 3) rules/map 매핑
        mapped_json, _meta, _rule, _src = _resolve_event_animation_entry(seq, mapping_payload)
        if not mapped_json:
            return 0.0

        # 4) JSON 파싱 + 총 길이 추정
        pth = _normalize_json_path(mapped_json)
        if not pth.exists():
            return 0.0
        parsed_steps = json.loads(pth.read_text(encoding="utf-8"))
        if not isinstance(parsed_steps, list):
            return 0.0
        est = _estimate_sequence_total_duration_sec_for_log(parsed_steps)
        return max(0.0, float(est)) if isinstance(est, (float, int)) else 0.0
    except Exception:
        return 0.0


def build_control_window(ext: Any) -> None:
    """TBS 제어창을 만들고 ext에 위젯/모델 참조를 저장."""
    # destroy()가 실패하거나(Kit 이벤트/프레임 타이밍), 핫리로드로 ext 인스턴스가 바뀌면
    # 이전 창이 화면에 남은 채로 새 창이 생성되어 UI가 겹쳐 보일 수 있다.
    # 1) ext 참조 기준 중복 생성 방지
    if getattr(ext, "_control_window", None) is not None:
        return
    # 2) 워크스페이스에 남아있는 동명 창이 있으면 선제 제거(핫리로드/비정상 destroy 대비)
    try:
        ws = getattr(ui, "Workspace", None)
        if ws is not None and hasattr(ws, "get_window"):
            old = ws.get_window("TBS 제어창")
            if old is not None:
                try:
                    old.destroy()
                except Exception:
                    try:
                        old.visible = False
                    except Exception:
                        pass
    except Exception:
        pass

    ext._xml_from_port_model = ui.SimpleIntModel(1)
    ext._xml_to_port_model = ui.SimpleIntModel(6)
    ext._xml_port_id_model = ui.SimpleIntModel(1)
    ext._last_generated_xml = ""
    ext._priority_prefix_model = ui.SimpleStringModel(DEFAULT_PRIORITY_NAME_PREFIX)
    ext._sim_lot_count_model = ui.SimpleIntModel(6)
    ext._sim_lot_spawn_min_model = ui.SimpleFloatModel(15.0)
    ext._sim_lot_spawn_max_model = ui.SimpleFloatModel(40.0)
    ext._sim_pickup_evt_min_model = ui.SimpleFloatModel(50.0)
    ext._sim_pickup_evt_max_model = ui.SimpleFloatModel(70.0)
    ext._sim_speed_model = ui.SimpleFloatModel(1.0)
    ext._sim_log_interval_model = ui.SimpleFloatModel(0.0)
    ext._sim_confirm_each_step_model = ui.SimpleBoolModel(False)
    ext._sim_oht_bp1_min_model = ui.SimpleFloatModel(5.0)
    ext._sim_oht_bp1_max_model = ui.SimpleFloatModel(10.0)
    ext._sim_bp1_bp_min_model = ui.SimpleFloatModel(5.0)
    ext._sim_bp1_bp_max_model = ui.SimpleFloatModel(10.0)
    ext._sim_bp_ep_min_model = ui.SimpleFloatModel(5.0)
    ext._sim_bp_ep_max_model = ui.SimpleFloatModel(10.0)
    ext._sim_ep_oht_min_model = ui.SimpleFloatModel(5.0)
    ext._sim_ep_oht_max_model = ui.SimpleFloatModel(10.0)
    ext._sim_ep_count_combo = None
    ext._sim_init_bp1_model = ui.SimpleBoolModel(False)
    ext._sim_init_bp2_model = ui.SimpleBoolModel(False)
    ext._sim_init_bp3_model = ui.SimpleBoolModel(False)
    ext._sim_init_bp4_model = ui.SimpleBoolModel(False)
    ext._sim_init_ep1_model = ui.SimpleBoolModel(False)
    ext._sim_init_ep2_model = ui.SimpleBoolModel(False)
    ext._sim_init_ep3_model = ui.SimpleBoolModel(False)
    ext._sim_init_ep3_row = None
    ext._sim_log_text = ui.SimpleStringModel("[SIM] 대기 중")
    ext._sim_history_text = ui.SimpleStringModel("[SIM] 대기 중")
    ext._sim_progress_text = ui.SimpleStringModel("[진행현황] 없음")
    ext._sim_port_state_text = ui.SimpleStringModel("[포트상태] 대기 중")
    # 요약(애니 실행이력 창): "스토리 1개 + 그 이후 연결된 JSON 목록" 블록을 유지
    # block: {"story": str, "anims": List[str]}
    ext._sim_recent_story_blocks = []
    # (요청으로 제거) 생성/회수 대기 토큰 표시 기능 비활성화
    # (요청으로 제거) 포트상태 좌/우 점 표시 기능은 비활성화
    ext._sim_progress_rows = {}
    ext._sim_progress_history = []
    ext._sim_progress_start_times = {}
    # 진행현황 RUNNING 라인 디듀프: percent/elapsed/total이 같으면 UI 갱신 스킵
    ext._sim_progress_last_key = {}
    ext._sim_engine = None
    ext._sim_update_sub = None
    ext._sim_thread = None
    ext._sim_thread_stop = None
    ext._sim_log_queue = None
    ext._sim_log_ui_sub = None
    ext._sim_log_view_combo = None
    ext._sim_progress_frame = None
    ext._sim_history_frame = None
    ext._sim_anim_history_frame = None
    ext._sim_port_state_frame = None
    ext._sim_port_state_header_label = None
    ext._sim_port_cells = {}
    ext._sim_port_cell_boxes = {}
    ext._sim_port_ep3_cell = None
    ext._sim_port_ep3_cell_container = None
    ext._sim_progress_label = None
    ext._sim_history_label = None
    ext._sim_anim_history_label = None
    ext._sim_port_state_label = None
    ext._sim_runner = SequenceRunner()
    ext._sim_anim_active = {}
    ext._sim_anim_pending = []
    # 애니메이션 재생 중 sim tick을 잠시 멈추기 위한 플래그
    ext._sim_tick_pause_event = threading.Event()
    # 이벤트 확인창(공정확인) 표시 중 sim tick을 잠시 멈추기 위한 플래그
    ext._sim_gate_pause_event = threading.Event()
    # fail-safe: 예상 애니 길이만큼은 최소 pause 유지 (monotonic timestamp)
    ext._sim_tick_pause_until_wall = None
    ext._sim_gate_dialog = None

    ext._control_window = ui.Window("TBS 제어창", width=800, height=840)
    with ext._control_window.frame:
        with ui.ScrollingFrame(
            height=ui.Fraction(1.0),
            style={"ScrollingFrame": {"padding": 4, "margin": 0}},
        ):
            with ui.VStack(spacing=0):
                ext._kit_chrome_hide_model = ui.SimpleBoolModel(KIT_CHROME_HIDE_DEFAULT_ON_LAUNCH)

                def _on_kit_chrome_toggle(model):
                    try:
                        apply_kit_chrome_hidden(ext, bool(model.as_bool))
                    except Exception:
                        pass

                ext._kit_chrome_hide_model.add_value_changed_fn(_on_kit_chrome_toggle)

                with ui.Frame(style={"background_color": 0xFF23262B}):
                    with ui.VStack(padding=8, spacing=8):
                        ui.Label("화면", height=24, style={"color": 0xFFDDDDDD})
                        with ui.HStack(spacing=8, height=28):
                            ui.Label(
                                "기본 메뉴·패널 숨기기 (3D 뷰·TBS·시퀀스 편집기 유지)",
                                width=0,
                                style={"color": 0xFFCCCCCC},
                            )
                            ui.CheckBox(
                                model=ext._kit_chrome_hide_model,
                                width=28,
                                style=CHECKBOX_WHITE_STYLE,
                            )
                ui.Spacer(height=6)
                with ui.Frame(style={"background_color": 0xFF23262B}):
                    with ui.VStack(padding=8, spacing=8):
                        ui.Label("USD Load", height=24, style={"color": 0xFFDDDDDD})
                        build_load_ui_into_stack(ext)
                ui.Spacer(height=6)
                with ui.Frame(style={"background_color": 0xFF23262B}):
                    # 콤보에 과도한 width 지정 시 Kit에서 다음 구역과 겹침이 발생할 수 있어 세로 스택만 사용
                    with ui.VStack(padding=8, spacing=8):
                        ui.Label("XML 제너레이터 생성기", height=24, style={"color": 0xFFDDDDDD})
                        ext._xml_seq_combo = ui.ComboBox(
                            0,
                            xml_generator.SEQ_READYTOLOAD,
                            xml_generator.SEQ_ARRIVED,
                            xml_generator.SEQ_MOVE_TRANSFERING,
                            xml_generator.SEQ_MOVE,
                            xml_generator.SEQ_MOVE_REQ,
                            xml_generator.SEQ_READYTOUNLOAD,
                            xml_generator.SEQ_REMOVED,
                        )
                        ext._xml_seq_combo.model.add_item_changed_fn(lambda m, *a: on_xml_seq_changed(ext))
                        with ui.HStack(spacing=8, height=28):
                            ui.Button("OK", width=72, height=28, clicked_fn=lambda: on_xml_ok_clicked(ext))
                            ui.Button("제너레이터 실행(역파싱)", height=28, clicked_fn=lambda: on_xml_run_clicked(ext))
                        ext._xml_ab_inputs_frame = ui.HStack(spacing=8, height=28)
                        with ext._xml_ab_inputs_frame:
                            ui.Label("FROM_PORT_ID", width=110, height=28)
                            ui.IntField(model=ext._xml_from_port_model, width=60, height=28)
                            ui.Label("TO_PORT_ID", width=90, height=28)
                            ui.IntField(model=ext._xml_to_port_model, width=60, height=28)
                        ext._xml_ab_inputs_frame.visible = True

                        ext._xml_port_inputs_frame = ui.HStack(spacing=8, height=28)
                        with ext._xml_port_inputs_frame:
                            ui.Label("PORT_ID", width=110, height=28)
                            ui.IntField(model=ext._xml_port_id_model, width=60, height=28)
                        ext._xml_port_inputs_frame.visible = False
                        # 콤보 초기 선택값 기준으로 입력 필드 표시 상태 동기화
                        on_xml_seq_changed(ext)
                ui.Spacer(height=6)
                ui.Rectangle(height=2, style={"background_color": 0xFF3A3A3A})
                ui.Spacer(height=6)
                with ui.Frame(style={"background_color": 0xFF1E2530}):
                    with ui.VStack(padding=8, spacing=6):
                        ui.Label("시뮬레이션 (simpy)", height=24, style={"color": 0xFFDDDDDD})
                        with ui.HStack(spacing=8, height=28):
                            ui.Label("LOT 수", width=80)
                            ui.IntField(model=ext._sim_lot_count_model, width=80)
                            ui.Label("EP 개수", width=55)
                            ext._sim_ep_count_combo = ui.ComboBox(0, "2", "3")
                            ext._sim_ep_count_combo.model.add_item_changed_fn(lambda m, *a: on_sim_ep_count_changed(ext))
                        with ui.HStack(spacing=8, height=28):
                            ui.Label("LOT생성간격", width=100)
                            ui.FloatField(model=ext._sim_lot_spawn_min_model, width=65)
                            ui.Label("~", width=10)
                            ui.FloatField(model=ext._sim_lot_spawn_max_model, width=65)
                            ui.Label("회수간격", width=60)
                            ui.FloatField(model=ext._sim_pickup_evt_min_model, width=55)
                            ui.Label("~", width=10)
                            ui.FloatField(model=ext._sim_pickup_evt_max_model, width=55)
                        ui.Label("초기 LOT 적재 포트 (체크 시 시작 시점에 FULL)", height=20)
                        with ui.HStack(spacing=8, height=26):
                            ui.Label("BP1", width=30); ui.CheckBox(model=ext._sim_init_bp1_model, width=30, style=CHECKBOX_WHITE_STYLE)
                            ui.Label("BP2", width=30); ui.CheckBox(model=ext._sim_init_bp2_model, width=30, style=CHECKBOX_WHITE_STYLE)
                            ui.Label("BP3", width=30); ui.CheckBox(model=ext._sim_init_bp3_model, width=30, style=CHECKBOX_WHITE_STYLE)
                            ui.Label("BP4", width=30); ui.CheckBox(model=ext._sim_init_bp4_model, width=30, style=CHECKBOX_WHITE_STYLE)
                        with ui.HStack(spacing=8, height=26):
                            ui.Label("EP1", width=30); ui.CheckBox(model=ext._sim_init_ep1_model, width=30, style=CHECKBOX_WHITE_STYLE)
                            ui.Label("EP2", width=30); ui.CheckBox(model=ext._sim_init_ep2_model, width=30, style=CHECKBOX_WHITE_STYLE)
                            ext._sim_init_ep3_row = ui.HStack(spacing=8, height=26)
                            with ext._sim_init_ep3_row:
                                ui.Label("EP3", width=30); ui.CheckBox(model=ext._sim_init_ep3_model, width=30, style=CHECKBOX_WHITE_STYLE)
                        try:
                            ext._sim_init_ep3_model.add_value_changed_fn(lambda m: on_sim_ep_count_changed(ext))
                        except Exception:
                            pass
                        for mdl in (
                            ext._sim_init_bp1_model,
                            ext._sim_init_bp2_model,
                            ext._sim_init_bp3_model,
                            ext._sim_init_bp4_model,
                            ext._sim_init_ep1_model,
                            ext._sim_init_ep2_model,
                            ext._sim_init_ep3_model,
                        ):
                            try:
                                mdl.add_value_changed_fn(lambda m: _sync_ep3_port_cell_visibility(ext))
                            except Exception:
                                pass
                        on_sim_ep_count_changed(ext)
                        with ui.HStack(spacing=8, height=28):
                            ui.Label("OHT→BP/EP", width=100)
                            ui.FloatField(model=ext._sim_oht_bp1_min_model, width=70)
                            ui.Label("~", width=10)
                            ui.FloatField(model=ext._sim_oht_bp1_max_model, width=70)
                            ui.Label("BP1->BP", width=60)
                            ui.FloatField(model=ext._sim_bp1_bp_min_model, width=55)
                            ui.Label("~", width=10)
                            ui.FloatField(model=ext._sim_bp1_bp_max_model, width=55)
                        with ui.HStack(spacing=8, height=28):
                            ui.Label("BP->EP", width=80)
                            ui.FloatField(model=ext._sim_bp_ep_min_model, width=70)
                            ui.Label("~", width=10)
                            ui.FloatField(model=ext._sim_bp_ep_max_model, width=70)
                            ui.Label("EP->OHT", width=60)
                            ui.FloatField(model=ext._sim_ep_oht_min_model, width=55)
                            ui.Label("~", width=10)
                            ui.FloatField(model=ext._sim_ep_oht_max_model, width=55)
                        with ui.HStack(spacing=8, height=28):
                            ui.Label("시뮬 속도배율", width=100)
                            ui.FloatField(model=ext._sim_speed_model, width=80)
                            ui.Label("로그주기(s)", width=70)
                            ui.FloatField(model=ext._sim_log_interval_model, width=70)
                            ui.CheckBox(model=ext._sim_confirm_each_step_model, width=30, style=CHECKBOX_WHITE_STYLE)
                            ui.Label("각 공정 확인", width=80)
                            ui.Button("시작", width=80, clicked_fn=lambda: on_sim_start_clicked(ext))
                            ui.Button("정지", width=80, clicked_fn=lambda: on_sim_stop_clicked(ext))
                            ui.Button("리셋", width=80, clicked_fn=lambda: on_sim_reset_clicked(ext))
                        with ui.HStack(spacing=8, height=24):
                            ui.Label("표시모드", width=60)
                            ext._sim_log_view_combo = ui.ComboBox(0, "둘다", "진행현황", "이력로그")
                            ext._sim_log_view_combo.model.add_item_changed_fn(lambda m, *a: on_sim_log_view_changed(ext))
                            ui.Button("진행현황 복사", width=100, clicked_fn=lambda: on_copy_sim_progress(ext))
                        ext._sim_port_state_frame = ui.ScrollingFrame(height=120)
                        with ext._sim_port_state_frame:
                            with ui.VStack(spacing=4):
                                ext._sim_port_state_header_label = ui.Label("[포트상태] 대기 중", height=20, style={"color": 0xFFBFE7FF})
                                # 포트 표
                                with ui.VStack(spacing=4):
                                    with ui.HStack(spacing=4, height=24):
                                        with ui.ZStack(width=90, height=24):
                                            ext._sim_port_cell_boxes["BP2"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_cells["BP2"] = ui.Label("BP2:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                        with ui.ZStack(width=90, height=24):
                                            ext._sim_port_cell_boxes["BP3"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_cells["BP3"] = ui.Label("BP3:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                        with ui.ZStack(width=90, height=24):
                                            ext._sim_port_cell_boxes["BP4"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_cells["BP4"] = ui.Label("BP4:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                    with ui.HStack(spacing=4, height=24):
                                        with ui.ZStack(width=90, height=24):
                                            ext._sim_port_cell_boxes["BP1"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_cells["BP1"] = ui.Label("BP1:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                        with ui.ZStack(width=90, height=24):
                                            ext._sim_port_cell_boxes["EP1"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_cells["EP1"] = ui.Label("EP1:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                        with ui.ZStack(width=90, height=24):
                                            ext._sim_port_cell_boxes["EP2"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_cells["EP2"] = ui.Label("EP2:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                        ext._sim_port_ep3_cell_container = ui.ZStack(width=90, height=24)
                                        with ext._sim_port_ep3_cell_container:
                                            ext._sim_port_cell_boxes["EP3"] = ui.Rectangle(style={"background_color": 0xFF2A2F38, "border_color": 0xFF7B8799, "border_width": 1})
                                            ext._sim_port_ep3_cell = ui.Label("EP3:-", width=90, height=24, style={"color": 0xFFFFFFFF})
                                ext._sim_port_state_label = ui.Label("", word_wrap=False, width=0, height=0, visible=False)
                        # 포트 상태 UI 구성 이후 EP3 표시조건 즉시 동기화
                        on_sim_ep_count_changed(ext)
                        _sync_ep3_port_cell_visibility(ext)
                        ext._sim_progress_frame = ui.ScrollingFrame(height=120)
                        with ext._sim_progress_frame:
                            ext._sim_progress_label = ui.Label(
                                "", word_wrap=True, height=118, style={"color": 0xFFFFFFFF}
                            )
                            ext._sim_progress_label.text = ext._sim_progress_text.as_string
                        ext._sim_history_frame = ui.ScrollingFrame(height=140)
                        with ext._sim_history_frame:
                            ext._sim_history_label = ui.Label(
                                "", word_wrap=True, height=136, style={"color": 0xFFFFFFFF}
                            )
                            ext._sim_history_label.text = ext._sim_history_text.as_string
                        on_sim_log_view_changed(ext)
                ui.Spacer(height=6)
                ui.Rectangle(height=2, style={"background_color": 0xFF3A3A3A})
                ui.Spacer(height=8)
                ui.Label("우선 표시 이름 규칙 (접두사, 비우면 순서대로 표시)", height=20)
                ui.StringField(model=ext._priority_prefix_model, height=22)
                ui.Spacer(height=4)
                ui.Label("로드된 USD 내 장비 prim (드롭다운)", height=20)
                ui.Button("목록 새로고침", height=28, clicked_fn=lambda: on_refresh_prim_list(ext))
                ui.Spacer(height=4)
                with ui.ScrollingFrame(height=280, style={"ScrollingFrame": {"padding": 0, "margin": 0}}):
                    ext._object_list_frame = ui.VStack(spacing=4, alignment=ui.Alignment.LEFT_TOP)
    refresh_object_list(ext)


def on_xml_seq_changed(ext: Any) -> None:
    try:
        idx = ext._xml_seq_combo.model.get_item_value_model().as_int
    except Exception:
        idx = 0
    seqs = [
        xml_generator.SEQ_READYTOLOAD,
        xml_generator.SEQ_ARRIVED,
        xml_generator.SEQ_MOVE_TRANSFERING,
        xml_generator.SEQ_MOVE,
        xml_generator.SEQ_READYTOUNLOAD,
        xml_generator.SEQ_REMOVED,
    ]
    seq = seqs[idx] if 0 <= idx < len(seqs) else xml_generator.SEQ_READYTOLOAD
    ext._xml_ab_inputs_frame.visible = seq in xml_generator.FROM_TO_SEQS
    ext._xml_port_inputs_frame.visible = seq in xml_generator.PORT_ID_ONLY_SEQS


def on_xml_ok_clicked(ext: Any) -> None:
    try:
        idx = ext._xml_seq_combo.model.get_item_value_model().as_int
    except Exception:
        idx = 0
    seqs = [
        xml_generator.SEQ_READYTOLOAD,
        xml_generator.SEQ_ARRIVED,
        xml_generator.SEQ_MOVE_TRANSFERING,
        xml_generator.SEQ_MOVE,
        xml_generator.SEQ_READYTOUNLOAD,
        xml_generator.SEQ_REMOVED,
    ]
    seq = seqs[idx] if 0 <= idx < len(seqs) else xml_generator.SEQ_READYTOLOAD
    try:
        if seq in xml_generator.FROM_TO_SEQS:
            from_port = ext._xml_from_port_model.get_value_as_int()
            to_port = ext._xml_to_port_model.get_value_as_int()
            xml = xml_generator.build_xml_string(seq, from_port_id=from_port, to_port_id=to_port)
        else:
            port_id = ext._xml_port_id_model.get_value_as_int()
            xml = xml_generator.build_xml_string(seq, port_id=port_id)
        ext._last_generated_xml = xml
        print(xml, flush=True)
    except Exception as e:
        print(f"[morph.tbs_control_1][xml_generator] XML 생성 실패: {e}", flush=True)


def on_xml_run_clicked(ext: Any) -> None:
    xml_text = (ext._last_generated_xml or "").strip()
    if not xml_text:
        print("[morph.tbs_control_1][xml_generator] 저장된 XML이 없습니다. 먼저 OK로 XML을 생성하세요.", flush=True)
        return
    parsed = xml_generator.parse_xml_string(xml_text)
    if not parsed:
        print("[morph.tbs_control_1][xml_generator] XML 역파싱 실패.", flush=True)
        return
    lines = ["[XML PARSE RESULT]"]
    if parsed.get("action_desc"):
        lines.append("[ACTION]")
        lines.append(parsed.get("action_desc", ""))

    for k in (
        "sequence_name",
        "destination",
        "origination",
        "tid",
        "facility",
        "equipment_id",
        "port_id",
        "from_port_id",
        "to_port_id",
    ):
        lines.append(f"{k} = {parsed.get(k, '')}")
    print("\n".join(lines), flush=True)


def _append_sim_log(ext: Any, line: str) -> None:
    """UI 스레드 전용: 이력 로그 패널(_sim_history_*)에 줄 추가. 시뮬 스레드는 post_sim_history_line 사용."""
    msg = _format_history_line((line or "").strip())
    if not msg:
        return
    prev = ext._sim_history_text.as_string if getattr(ext, "_sim_history_text", None) else ""
    merged = f"{prev}\n{msg}".strip() if prev else msg
    # 화면이 너무 길어지지 않게 최근 200줄만 유지
    rows = merged.splitlines()
    if len(rows) > 200:
        merged = "\n".join(rows[-200:])
    if getattr(ext, "_sim_history_text", None):
        ext._sim_history_text.set_value(merged)
    if getattr(ext, "_sim_history_label", None) is not None:
        ext._sim_history_label.text = merged


def _format_history_line(line: str) -> str:
    """
    이력로그 가독성을 위해 핵심 토큰을 앞에 배치하고 태그를 단순화한다.
    원문 정보는 유지하되 읽기만 쉽게 만든다.
    """
    s = (line or "").strip()
    if not s:
        return ""
    # 자주 보이는 태그를 직관적인 짧은 라벨로 치환
    tag_map = {
        "[STORY]": "[스토리]",
        "[MOVE]": "[이송]",
        "[ARRIVED]": "[도착]",
        "[PROCESS]": "[공정]",
        "[READYTOLOAD]": "[대기준비]",
        "[READYTOUNLOAD]": "[반출준비]",
        "[REMOVED]": "[반출완료]",
        "[INPUT]": "[투입]",
        "[WAIT]": "[대기]",
        "[SUMMARY LOT]": "[LOT 요약]",
        "[SUMMARY]": "[요약]",
    }
    for old, new in tag_map.items():
        if old in s:
            s = s.replace(old, new)
    # 시뮬 이벤트 원문은 너무 길어져서 핵심만 요약
    if "[SIM EVENT" in s and "seq=" in s:
        try:
            part = s.split("] ", 1)
            head = part[0] + "]" if len(part) == 2 else "[SIM EVENT]"
            body = part[1] if len(part) == 2 else s
            seq = ""
            lot = ""
            fr = ""
            to = ""
            for tok in body.split():
                if tok.startswith("seq="):
                    seq = tok[4:]
                elif tok.startswith("lot="):
                    lot = tok[4:]
                elif tok.startswith("from="):
                    fr = tok[5:]
                elif tok.startswith("to="):
                    to = tok[3:]
            route = f"{fr}->{to}" if fr and to else (to or fr or "-")
            return _with_history_color_icon(f"{head} [이벤트] seq={seq} lot={lot} route={route}")
        except Exception:
            return _with_history_color_icon(s)
    return _with_history_color_icon(s)


def _with_history_color_icon(s: str) -> str:
    """
    단일 Label 제약에서 줄 단위 강조를 위해 색상 아이콘을 앞에 붙인다.
    🟥 오류/실패, 🟨 대기/주의, 🟩 완료/성공, 🟦 이벤트/진행, ⬜ 일반
    """
    t = (s or "").upper()
    if any(k in t for k in ("실패", "ERROR", "EXCEPTION", "예외", "파싱실패")):
        return f"🟥 {s}"
    if any(k in t for k in ("대기", "[WAIT]", "파일없음", "매핑없음", "주의")):
        return f"🟨 {s}"
    if any(k in t for k in ("완료", "DONE", "저장 완료", "실행시작", "실행준비완료")):
        return f"🟩 {s}"
    if any(k in t for k in ("이벤트", "MOVE", "ARRIVED", "PROCESS", "STORY", "투입", "이송", "공정", "도착")):
        return f"🟦 {s}"
    return f"⬜ {s}"


def _append_anim_history_log(ext: Any, line: str) -> None:
    """애니메이션 실행이력 패널 제거됨. 호출은 호환을 위해 유지한다."""
    return


def _render_pending_dots(ext: Any) -> None:
    """(요청으로 제거) 점 표시 기능 비활성화."""
    return


def _port_cell_text(occ: Dict[str, Any], port: str) -> str:
    v = str(occ.get(port, "-")).strip()
    if not v or v.upper() in ("EMPTY", "-", "NONE"):
        return "-"
    if v.upper() == "FULL":
        return "FULL"
    return v


def _compact_cell_value(v: str, max_len: int = 10) -> str:
    s = (v or "-").strip()
    if len(s) <= max_len:
        return s
    return s[: max(1, max_len - 1)] + "..."


def _sync_ep3_port_cell_visibility(ext: Any) -> None:
    container = getattr(ext, "_sim_port_ep3_cell_container", None)
    if container is None:
        return
    try:
        ep_idx = ext._sim_ep_count_combo.model.get_item_value_model().as_int
    except Exception:
        ep_idx = 0
    checked = False
    if getattr(ext, "_sim_init_ep3_model", None) is not None:
        try:
            checked = bool(ext._sim_init_ep3_model.get_value_as_bool())
        except Exception:
            checked = False
    # 일부 환경에서 체크 이벤트 반영이 지연되는 문제를 피하기 위해
    # EP 개수=3이면 EP3 칸은 항상 보이게 유지하고, 체크 여부는 초기 적재 로직에서 사용.
    container.visible = bool(ep_idx == 1)


def _set_port_box_style(ext: Any, port: str, value: str) -> None:
    box = (getattr(ext, "_sim_port_cell_boxes", {}) or {}).get(port)
    if box is None:
        return
    v = (value or "").strip().upper()
    if not v or v in ("-", "EMPTY", "NONE"):
        fill = 0xFF2A2F38
    elif v == "FULL":
        fill = 0xFF6B5B2A
    else:
        fill = 0xFF1F4A36
    try:
        box.style = {"background_color": fill, "border_color": 0xFF7B8799, "border_width": 1}
    except Exception:
        pass


def _update_port_occupancy_panel(ext: Any, occ: Dict[str, Any], sim_time: str = "") -> None:
    if not isinstance(occ, dict):
        return
    bp2 = _port_cell_text(occ, "BP2")
    bp3 = _port_cell_text(occ, "BP3")
    bp4 = _port_cell_text(occ, "BP4")
    bp1 = _port_cell_text(occ, "BP1")
    ep1 = _port_cell_text(occ, "EP1")
    ep2 = _port_cell_text(occ, "EP2")
    ep3 = _port_cell_text(occ, "EP3")
    t = str(sim_time).strip()
    head = f"[포트상태 t={t}]" if t else "[포트상태]"
    if getattr(ext, "_sim_port_state_header_label", None) is not None:
        ext._sim_port_state_header_label.text = head
    _sync_ep3_port_cell_visibility(ext)
    cells = getattr(ext, "_sim_port_cells", {}) or {}
    if "BP2" in cells:
        cells["BP2"].text = f"BP2:{_compact_cell_value(bp2)}"
        _set_port_box_style(ext, "BP2", bp2)
    if "BP3" in cells:
        cells["BP3"].text = f"BP3:{_compact_cell_value(bp3)}"
        _set_port_box_style(ext, "BP3", bp3)
    if "BP4" in cells:
        cells["BP4"].text = f"BP4:{_compact_cell_value(bp4)}"
        _set_port_box_style(ext, "BP4", bp4)
    if "BP1" in cells:
        cells["BP1"].text = f"BP1:{_compact_cell_value(bp1)}"
        _set_port_box_style(ext, "BP1", bp1)
    if "EP1" in cells:
        cells["EP1"].text = f"EP1:{_compact_cell_value(ep1)}"
        _set_port_box_style(ext, "EP1", ep1)
    if "EP2" in cells:
        cells["EP2"].text = f"EP2:{_compact_cell_value(ep2)}"
        _set_port_box_style(ext, "EP2", ep2)
    ep3_cell = getattr(ext, "_sim_port_ep3_cell", None)
    if ep3_cell is not None:
        ep3_cell.text = f"EP3:{_compact_cell_value(ep3)}"
        _set_port_box_style(ext, "EP3", ep3)


def _enqueue_sim_log(ext: Any, line: str) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.HISTORY_LINE, (line or "").strip()))
    except Exception:
        pass


def post_sim_history_line(ext: Any, line: str) -> None:
    """시뮬 워커 스레드에서 호출: 스토리/상태 텍스트를 '이력 로그' 패널로 보낸다."""
    _enqueue_sim_log(ext, line)


def _enqueue_anim_event(ext: Any, payload: Dict[str, str]) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.ANIM_EVENT, dict(payload or {})))
    except Exception:
        pass


def post_sim_anim_event(ext: Any, payload: Dict[str, str]) -> None:
    """시뮬 워커 스레드에서 호출: 애니메이션 파이프라인(포트 패널 + 애니 이력)으로 이벤트를 넘긴다."""
    _enqueue_anim_event(ext, payload)


def _enqueue_control_action(ext: Any, action: str) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.ACTION, action))
    except Exception:
        pass


def _enqueue_gate_request(ext: Any, payload: Dict[str, Any]) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.GATE, dict(payload or {})))
    except Exception:
        pass


def _show_sim_gate_dialog(ext: Any, payload: Dict[str, Any]) -> None:
    # 공정확인 모드에서는 "확인 전까지 완전 정지"가 목표이므로,
    # 확인창이 이미 떠 있으면 새 창으로 교체하지 않고(=pause가 풀리는 부작용 방지) 그냥 대기시킨다.
    if getattr(ext, "_sim_gate_dialog", None) is not None:
        return
    title = str(payload.get("title", "공정 확인"))
    msg = str(payload.get("message", "다음 공정을 진행할까요?"))
    done = payload.get("_done_event", None)
    g_raw = str(payload.get("gate_seq_raw", "")).strip()
    g_can = str(payload.get("gate_seq_canonical", "")).strip()
    g_xml = str(payload.get("gate_xml_sequence_name", "")).strip()
    win_suffix = f" [{g_raw}]" if g_raw else ""
    ext._sim_gate_dialog = ui.Window(f"[SIM 확인] {title}{win_suffix}", width=580, height=400)
    with ext._sim_gate_dialog.frame:
        with ui.VStack(spacing=8, padding=10):
            with ui.Frame(style={"background_color": 0xFF2A3140, "border_width": 1, "border_color": 0xFF5A6A80}):
                with ui.VStack(spacing=4, padding=8):
                    ui.Label("이벤트 (sequence_name)", height=22, style={"color": 0xFF8EC8FF})
                    if g_raw and g_can and g_raw == g_can:
                        seq_line = f"sequence_name: {g_raw}"
                    elif g_raw or g_can:
                        seq_line = f"시뮬 seq: {g_raw or '-'}  → 규격/별칭: {g_can or '-'}"
                    else:
                        seq_line = "sequence_name: -"
                    ui.Label(seq_line, word_wrap=True, height=36)
                    if g_xml:
                        ui.Label(f"XML SEQUENCE_NAME: {g_xml}", height=22, style={"color": 0xFFC8E0FF})
            with ui.ScrollingFrame(height=240):
                with ui.VStack(spacing=4):
                    ui.Label(msg, word_wrap=True, height=200)
            with ui.HStack(spacing=8, height=30):
                ui.Button("확인", width=80, clicked_fn=lambda: _close_sim_gate_dialog(ext, done))


def _close_sim_gate_dialog(ext: Any, done_event: Any) -> None:
    w = getattr(ext, "_sim_gate_dialog", None)
    if w is not None:
        try:
            w.visible = False
            # 이벤트/드로우 중 destroy 호출 금지: 다음 프레임으로 지연
            def _defer_destroy(_e=None):
                try:
                    w.destroy()
                except Exception:
                    pass
            try:
                app.get_app().get_post_update_event_stream().create_subscription_to_pop(
                    _defer_destroy,
                    name="morph.tbs_control_1:sim_gate_destroy",
                )
            except Exception:
                pass
        except Exception:
            pass
    ext._sim_gate_dialog = None
    # 이벤트 확인창이 닫히면 gate pause 해제 (애니 pause는 별도 이벤트로 유지)
    try:
        gp = getattr(ext, "_sim_gate_pause_event", None)
        if gp is not None:
            gp.clear()
    except Exception:
        pass
    try:
        if done_event is not None:
            done_event.set()
    except Exception:
        pass


def _enqueue_sim_progress(ext: Any, payload: Dict[str, str]) -> None:
    q = getattr(ext, "_sim_log_queue", None)
    if q is None:
        return
    try:
        q.put_nowait((SimUiQueueKind.PROGRESS, dict(payload or {})))
    except Exception:
        pass


def post_sim_progress_update(ext: Any, payload: Dict[str, str]) -> None:
    """시뮬 워커 스레드에서 호출: 공정 진행률/상태를 '진행현황' 패널로 보낸다."""
    _enqueue_sim_progress(ext, payload)


def _sim_ui_sink_progress(ext: Any, payload: Dict[str, Any]) -> None:
    _update_sim_progress(ext, payload if isinstance(payload, dict) else {})


def _build_sim_gate_request_payload(ext: Any, p: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    공정확인(게이트) 창에 표시할 title/message를 구성해 큐 payload(dict)로 반환.

    목적:
    - 게이트 메시지 구성 로직을 한 함수로 모아, 유지보수 시 흐름 추적 비용을 줄인다.
    - 기능 동작(표시 순서/매핑/타이머/XML)은 기존과 동일하게 유지한다.
    """
    try:
        seq_raw = str(p.get("seq", "") or "")
        seq_can = SIM_SEQ_ALIAS.get(seq_raw.strip(), seq_raw.strip()) if seq_raw else ""
        lot = str(p.get("lot_id", "") or "")
        lot_seq = str(p.get("lot_seq", "") or "")
        foup_id = str(p.get("foup_id", "") or "")
        fr = str(p.get("from_port_id", "") or "")
        to = str(p.get("to_port_id", "") or "")
        port = str(p.get("port_id", "") or "")
        t = str(p.get("sim_time", "") or "")
        title = f"EVENT t={t}" if t else "EVENT"

        # 공정확인 창 메시지 구성(요구사항 순서):
        # 1) 발생 이벤트명  2) 이벤트 동작 설명  3) 연계 애니 파일/존재/빈파일/불필요
        # 4) TIMER(생성/회수)  5) XML 표시
        lines: List[str] = []
        lines.append(f"[EVENT] sequence_name={seq_can or '-'} (raw={seq_raw or '-'})")
        lines.append(
            f"[EVENT] lot={lot or '-'}"
            + (f" (seq={lot_seq})" if lot_seq else "")
            + (f" foup={foup_id}" if foup_id else "")
            + f" | from={fr or '-'} to={to or '-'} port={port or '-'}"
        )

        xml_text = ""
        seq_for_mapping = seq_can
        parsed: Dict[str, Any] = {}
        try:
            if seq_can in xml_generator.FROM_TO_SEQS:
                xml_text = xml_generator.build_xml_string(
                    seq_can,
                    from_port_id=_parse_port_num(fr, 1),
                    to_port_id=_parse_port_num(to, 1),
                )
            else:
                xml_text = xml_generator.build_xml_string(seq_can, port_id=_parse_port_num(port, 1))
            parsed = xml_generator.parse_xml_string(xml_text) or {}
            parsed_seq = str(parsed.get("sequence_name", "") or "").strip().upper()
            if parsed_seq:
                seq_for_mapping = parsed_seq
        except Exception:
            xml_text = ""
            parsed = {}

        # 2) 이벤트 동작 설명
        action_desc = str(parsed.get("action_desc", "") or "").strip() if isinstance(parsed, dict) else ""
        if action_desc:
            lines.append(f"[ACTION] {action_desc}")
        elif seq_for_mapping:
            lines.append(f"[ACTION] (설명 없음) seq={seq_for_mapping}")
        else:
            lines.append("[ACTION] (설명 없음)")

        # 3) 연계된 애니메이션 파일/존재여부/비어있는 파일 여부
        try:
            seq_u = str(seq_can or "").strip().upper()
            is_anim_event = seq_u in (
                str(xml_generator.SEQ_ARRIVED).strip().upper(),
                str(xml_generator.SEQ_MOVE_TRANSFERING).strip().upper(),
                str(xml_generator.SEQ_MOVE_REQ).strip().upper(),
                str(xml_generator.SEQ_REMOVED).strip().upper(),
            )
        except Exception:
            is_anim_event = False

        if not is_anim_event:
            lines.append("[ANIM] 이 이벤트는 애니메이션이 필요없는 이벤트입니다.")
        else:
            try:
                mapping_payload = dict(p or {})
                mapping_payload["seq"] = seq_for_mapping
                if parsed:
                    mapping_payload["from_port_id"] = _normalize_port_text_from_xml(str(parsed.get("from_port_id", "") or ""), fr)
                    mapping_payload["to_port_id"] = _normalize_port_text_from_xml(str(parsed.get("to_port_id", "") or ""), to)
                    mapping_payload["port_id"] = _normalize_port_text_from_xml(str(parsed.get("port_id", "") or ""), port)
                mapped_json, _meta, rule_name, source_name = _resolve_event_animation_entry(seq_for_mapping, mapping_payload)
                if not mapped_json:
                    lines.append(f"[ANIM] 매핑 없음 (event={seq_for_mapping})")
                else:
                    jp = _normalize_json_path(mapped_json)
                    exists_txt = "존재" if jp.exists() else "없음"
                    empty_txt = ""
                    if jp.exists():
                        try:
                            raw = json.loads(jp.read_text(encoding="utf-8"))
                            if isinstance(raw, list) and len(raw) == 0:
                                empty_txt = " / EMPTY(빈 파일)"
                        except Exception:
                            empty_txt = ""
                    lines.append(
                        f"[ANIM] file={jp.name} ({exists_txt}{empty_txt}) | source={source_name or '-'} rule={rule_name or '-'}"
                    )
            except Exception as e:
                lines.append(f"[ANIM] 매핑 확인 실패: {e}")

        # 4) TIMER
        try:
            sim = getattr(ext, "_sim_engine", None)
            now_t = float(p.get("sim_time", "0.0") or 0.0)
            spawn_at = getattr(sim, "_next_spawn_at", None) if sim is not None else None
            pickup_at = getattr(sim, "_next_pickup_at", None) if sim is not None else None
            lines_t: List[str] = []
            if isinstance(spawn_at, (int, float)):
                lines_t.append(f"다음 생성까지: {max(0.0, float(spawn_at) - now_t):.2f}s (sim)")
            if isinstance(pickup_at, (int, float)):
                lines_t.append(f"다음 회수티켓까지: {max(0.0, float(pickup_at) - now_t):.2f}s (sim)")
            if lines_t:
                lines.append("")
                lines.append("TIMER:")
                lines.extend(lines_t)
        except Exception:
            pass

        # 5) XML
        if xml_text:
            lines.append("")
            lines.append("XML:")
            lines.append(xml_text)

        message = "\n".join([ln for ln in lines if ln is not None])
        return {
            "title": title,
            "message": message,
            "_done_event": threading.Event(),
            "gate_seq_raw": seq_raw,
            "gate_seq_canonical": seq_can,
            "gate_xml_sequence_name": "",
        }
    except Exception:
        return None


def _sim_ui_sink_anim_event(ext: Any, payload: Dict[str, Any], panel_mode: SimLogPanelMode) -> None:
    p = payload if isinstance(payload, dict) else {}
    occ = p.get("ports_occupancy", {})
    if not isinstance(occ, dict):
        occ = {}
    try:
        apply_port_lot_prim_visibility(occ)
    except Exception:
        pass
    _update_port_occupancy_panel(ext, occ, str(p.get("sim_time", "")))
    # 포트상태 갱신 전용 이벤트: 목록에 없는 내부 이벤트이므로 애니/공정확인창을 띄우지 않는다.
    try:
        if str(p.get("seq", "") or "").strip().upper() == "PORT_OCC_REFRESH":
            return
    except Exception:
        pass
    # 포트상태 좌/우 점(●) 카운터:
    # - READYTOLOAD 발생 시(생성 이벤트) 좌측 초록 ● +1
    # - READYTOUNLOAD 발생 시(회수 요청) 우측 빨강 ● +1
    # - 실제 감소는 애니 완료 시점(ARRIVED(OHT->*) 완료 / REMOVED 완료)에서 수행
    # (요청으로 제거) 포트상태 좌/우 점 표시 기능 비활성화
    verbose = panel_mode != SimLogPanelMode.PROGRESS_ONLY
    handle_sim_event_for_animation(ext, p, verbose=verbose)


def _sim_ui_sink_history_line(ext: Any, line: str, panel_mode: SimLogPanelMode) -> None:
    if not line:
        return
    if panel_mode == SimLogPanelMode.PROGRESS_ONLY:
        return
    _append_sim_log(ext, line)


def _sim_ui_sink_action(ext: Any, payload: Any) -> None:
    if str(payload) == SimUiControlAction.EXPORT_XLSX.value:
        _export_sim_logs_to_xlsx(ext)


def _sim_ui_sink_gate(ext: Any, payload: Dict[str, Any]) -> None:
    _show_sim_gate_dialog(ext, payload if isinstance(payload, dict) else {})


def _coerce_sim_ui_queue_kind(kind: Any) -> str:
    """
    큐에서 꺼낸 kind가 SimUiQueueKind 멤버일 때 str(kind)는 'SimUiQueueKind.GATE'처럼
    값이 아니라 멤버 이름이 되어 라우팅이 깨진다. 항상 실제 큐 문자열 값으로 맞춘다.
    """
    if isinstance(kind, SimUiQueueKind):
        return kind.value
    return str(kind)


def _dispatch_sim_ui_queue_item(ext: Any, kind: str, payload: Any, panel_mode: SimLogPanelMode) -> None:
    if kind == SimUiQueueKind.PROGRESS.value:
        _sim_ui_sink_progress(ext, payload if isinstance(payload, dict) else {})
    elif kind == SimUiQueueKind.ANIM_EVENT.value:
        _sim_ui_sink_anim_event(ext, payload if isinstance(payload, dict) else {}, panel_mode)
    elif kind == SimUiQueueKind.ACTION.value:
        _sim_ui_sink_action(ext, payload)
    elif kind == SimUiQueueKind.GATE.value:
        _sim_ui_sink_gate(ext, payload if isinstance(payload, dict) else {})
    elif kind == SimUiQueueKind.HISTORY_LINE.value:
        line = payload if isinstance(payload, str) else str(payload)
        _sim_ui_sink_history_line(ext, line, panel_mode)
    else:
        line = payload if isinstance(payload, str) else str(payload)
        _sim_ui_sink_history_line(ext, line, panel_mode)


def _drain_sim_log_queue(ext: Any) -> None:
    try:
        q = getattr(ext, "_sim_log_queue", None)
        if q is None:
            return
        try:
            view_idx = ext._sim_log_view_combo.model.get_item_value_model().as_int
        except Exception:
            view_idx = 0
        if view_idx > int(SimLogPanelMode.HISTORY_ONLY):
            view_idx = int(SimLogPanelMode.ALL)
        try:
            panel_mode = SimLogPanelMode(int(view_idx))
        except Exception:
            panel_mode = SimLogPanelMode.ALL
        count = 0
        # 중요: UI 프레임 1회당 처리량 상한.
        # 큐가 많아도 렌더링 starvation을 막기 위해 200개까지만 드레인한다.
        while count < 200:
            try:
                item = q.get_nowait()
            except Exception:
                break
            kind, payload = (
                item if isinstance(item, tuple) and len(item) == 2 else (SimUiQueueKind.HISTORY_LINE.value, item)
            )
            _dispatch_sim_ui_queue_item(ext, _coerce_sim_ui_queue_kind(kind), payload, panel_mode)
            count += 1

            # 공정확인 체크 + gate pause 상태면, "확인창 1개를 띄울 때까지만" 처리하고 멈춘다.
            # (gate pause를 너무 이르게 걸어도 UI가 1개 이벤트를 처리해 창을 띄울 수 있어야 한다)
            try:
                confirm_each = bool(
                    getattr(ext, "_sim_confirm_each_step_model", None) is not None
                    and ext._sim_confirm_each_step_model.get_value_as_bool()
                )
            except Exception:
                confirm_each = False
            if confirm_each:
                try:
                    gp = getattr(ext, "_sim_gate_pause_event", None)
                    if gp is not None and gp.is_set() and getattr(ext, "_sim_gate_dialog", None) is not None:
                        break
                except Exception:
                    break
    except Exception as e:
        # UI 드레인 예외가 발생해도 구독이 끊기지 않도록 보호
        print(f"[SIM UI] 로그 드레인 예외: {e}", flush=True)


def _sim_anim_status_key(ext: Any) -> Tuple[bool, str, int, str]:
    """진행 패널 중복 스킵용: 재생 여부·현재 파일·대기 큐·다음 파일."""
    runner = getattr(ext, "_sim_runner", None)
    try:
        running = bool(runner is not None and runner.is_running())
    except Exception:
        running = False
    active = getattr(ext, "_sim_anim_active", None) or {}
    cur_file = str(active.get("file", "") or "").strip() if isinstance(active, dict) else ""
    pend = getattr(ext, "_sim_anim_pending", None)
    plist = pend if isinstance(pend, list) else []
    q = len(plist)
    next_f = ""
    if plist and isinstance(plist[0], dict):
        next_f = str(plist[0].get("file", "") or "").strip()
    return (running, cur_file, q, next_f)


def _format_anim_status_footer(ext: Any) -> str:
    """진행현황 패널 하단: 현재 재생 JSON 파일·대기열."""
    running, cur_file, q, next_f = _sim_anim_status_key(ext)
    if running and cur_file:
        lines = [f"애니메이션 파일(재생 중): {cur_file}"]
        if q > 0 and next_f:
            lines.append(f"대기열: {q}건 (다음 {next_f})")
        return "\n".join(lines)
    if q > 0 and next_f:
        return "애니메이션: 대기 — 다음 " + next_f + (f" (큐 {q}건)" if q > 1 else "")
    return "애니메이션 파일: 재생 없음"


def _refresh_sim_progress_from_last(ext: Any) -> None:
    """애니 시작/종료 직후 마지막 공정 진행 payload로 패널만 다시 그린다."""
    lp = getattr(ext, "_sim_progress_last_payload", None)
    if isinstance(lp, dict):
        _update_sim_progress(ext, lp)


def _update_sim_progress(ext: Any, payload: Dict[str, str]) -> None:
    """진행현황 패널 갱신: 단일 창에 ‘현재 단계’만 표시(이벤트 로그 블록과 역할 분담)."""
    label = str(payload.get("label", "")).strip()
    if not label:
        return
    status = str(payload.get("status", "RUNNING"))
    percent = str(payload.get("percent", "0"))
    elapsed = str(payload.get("elapsed", "0.0"))
    total = str(payload.get("total", "0.0"))
    sim_time = str(payload.get("sim_time", "0.00"))
    detail = str(payload.get("detail", ""))
    event_seq = str(payload.get("event_seq") or payload.get("sequence_name") or "").strip()

    anim_key = _sim_anim_status_key(ext)
    if status == "RUNNING":
        try:
            last_key = getattr(ext, "_sim_progress_last_key", None)
            key = (str(percent), str(elapsed), str(total), str(status), label, anim_key)
            if isinstance(last_key, dict) and last_key.get("_single_panel") == key:
                return
            if isinstance(last_key, dict):
                last_key["_single_panel"] = key
        except Exception:
            pass
    else:
        try:
            last_key = getattr(ext, "_sim_progress_last_key", None)
            if isinstance(last_key, dict):
                last_key.pop("_single_panel", None)
        except Exception:
            pass

    head = "[진행현황] 단계 완료" if status == "DONE" else "[진행현황] 진행 중"
    ev_line = f"이벤트: {event_seq}\n" if event_seq else ""
    anim_footer = _format_anim_status_footer(ext)
    text = (
        f"{head} | t(sim)={sim_time}s\n"
        f"{ev_line}"
        f"{label}\n"
        f"진행률: {percent}% ({elapsed} / {total}s)\n"
        f"{detail}\n"
        f"---\n"
        f"{anim_footer}"
    )
    try:
        ext._sim_progress_last_payload = dict(payload)
    except Exception:
        ext._sim_progress_last_payload = payload
    ext._sim_progress_text.set_value(text)
    if getattr(ext, "_sim_progress_label", None) is not None:
        ext._sim_progress_label.text = text


def _on_sim_event(ext: Any, payload: Dict[str, str]) -> None:
    seq_raw = (payload.get("seq") or "").strip()
    if not seq_raw:
        return
    seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
    lot_id = payload.get("lot_id", "")
    sim_time = payload.get("sim_time", "")

    try:
        if seq in xml_generator.FROM_TO_SEQS:
            from_port = _parse_port_num(str(payload.get("from_port_id", "1")), 1)
            to_port = _parse_port_num(str(payload.get("to_port_id", "1")), 1)
            xml_text = xml_generator.build_xml_string(seq, from_port_id=from_port, to_port_id=to_port)
        else:
            port = _parse_port_num(str(payload.get("port_id", "1")), 1)
            xml_text = xml_generator.build_xml_string(seq, port_id=port)
        ext._last_generated_xml = xml_text
        parsed = xml_generator.parse_xml_string(xml_text) or {}
        story = f"[SIM EVENT t={sim_time}] seq={seq_raw}->{seq} lot={lot_id} port={payload.get('port_id','')} from={payload.get('from_port_id','')} to={payload.get('to_port_id','')}"
        if parsed.get("action_desc"):
            story += f" | action={parsed.get('action_desc')}"
        _append_sim_log(ext, story)
    except Exception as e:
        _append_sim_log(ext, f"[SIM EVENT] XML 생성/역파싱 실패: seq={seq}, err={e}")


def _parse_port_num(port_text: str, default_value: int = 1) -> int:
    """
    내부 포트 텍스트를 EAPEIS 포트 ID로 변환한다.
    매핑 규칙:
    - OHT -> 9 (MOVE FROM 가상 포트; EP/BP와 충돌 없음)
    - EP1/2/3 -> 1/2/3
    - BP1/2/3/4 -> 5/6/7/8
    """
    txt = (port_text or "").strip().upper()
    if not txt:
        return default_value
    if txt.startswith("OHT"):
        return 9
    if txt.startswith("EP"):
        try:
            n = int(txt.replace("EP", ""))
            if 1 <= n <= 3:
                return n
        except Exception:
            return default_value
    if txt.startswith("BP"):
        try:
            n = int(txt.replace("BP", ""))
            if 1 <= n <= 4:
                return 4 + n
        except Exception:
            return default_value
    if txt.startswith("PORT_"):
        txt = txt.replace("PORT_", "")
    try:
        return int(txt)
    except Exception:
        return default_value


def _port_kind(port_text: str) -> str:
    t = (port_text or "").strip().upper()
    if t.startswith("BP"):
        return "버퍼포트(BP)"
    if t.startswith("EP"):
        return "공정포트(EP)"
    if t.startswith("OHT"):
        return "이송장치(OHT)"
    return "미확인"


def _normalize_port_text_from_xml(parsed_val: str, original_text: str) -> str:
    """
    XML 역파싱 값(parsed_val)을 기준으로 포트 문자열을 표준화한다.
    - 원본이 BP/EP 접두사를 가지고 있으면 같은 접두사를 유지
    - 없으면 XML 숫자값 그대로 사용
    """
    p = (parsed_val or "").strip()
    if not p:
        return ""
    o = (original_text or "").strip().upper()
    try:
        n = int(p)
    except Exception:
        n = None
    if o.startswith("BP"):
        # XML ID 5~8은 BP1~4에 대응
        if n is not None and 5 <= n <= 8:
            return f"BP{n - 4}"
        return f"BP{p}"
    if o.startswith("EP"):
        # XML ID 1~3은 EP1~3에 대응
        if n is not None and 1 <= n <= 3:
            return f"EP{n}"
        return f"EP{p}"
    if o.startswith("OHT"):
        return "OHT"
    if n is not None:
        if 1 <= n <= 3:
            return f"EP{n}"
        if 5 <= n <= 8:
            return f"BP{n - 4}"
    return p


def handle_sim_event_for_animation(ext: Any, payload: Dict[str, str], verbose: bool = True) -> None:
    """
    시뮬레이션 이벤트 -> 애니메이션 실행 훅.
    현재는 분기별 로그만 출력하고, 추후 분기 내부에 실제 애니메이션 함수를 연결한다.
    """
    seq_raw = (payload.get("seq") or "").strip()
    if not seq_raw:
        return
    seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
    sim_time = payload.get("sim_time", "")
    lot_id = payload.get("lot_id", "")
    from_port_txt = str(payload.get("from_port_id", ""))
    to_port_txt = str(payload.get("to_port_id", ""))
    port_txt = str(payload.get("port_id", ""))
    from_kind = _port_kind(from_port_txt)
    to_kind = _port_kind(to_port_txt)
    port_kind = _port_kind(port_txt)

    if verbose:
        print(
            f"[ANIM HOOK t={sim_time}] 이벤트={seq_raw}->{seq} lot={lot_id} "
            f"port={port_txt}({port_kind}) from={from_port_txt}({from_kind}) to={to_port_txt}({to_kind})",
            flush=True,
        )

    # 요구사항: READYTOLOAD / READYTOUNLOAD 는 애니메이션을 절대 실행하지 않는다.
    # (rules/map에 남아있는 매핑이 있어도 무시)
    try:
        if str(seq).strip().upper() in (
            str(xml_generator.SEQ_READYTOLOAD).strip().upper(),
            str(xml_generator.SEQ_READYTOUNLOAD).strip().upper(),
        ):
            if verbose:
                print(f"[ANIM HOOK] no-anim event skip: {seq}", flush=True)
            return
    except Exception:
        pass

    # 주 실행 경로: 이벤트 -> XML 생성 -> 역파싱 -> 매핑 -> JSON 실행
    # 유지보수 규칙:
    # - rules/map 매칭 입력값은 "반드시" XML 역파싱 결과를 기준으로 표준화한다.
    # - 시뮬 payload 원본을 바로 매칭에 쓰지 않는다(포맷 드리프트 방지).
    parsed: Dict[str, Any] = {}
    xml_text = ""
    seq_for_mapping = seq
    try:
        if seq in xml_generator.FROM_TO_SEQS:
            from_port = _parse_port_num(from_port_txt, 1)
            to_port = _parse_port_num(to_port_txt, 1)
            xml_text = xml_generator.build_xml_string(seq, from_port_id=from_port, to_port_id=to_port)
        else:
            port = _parse_port_num(port_txt, 1)
            xml_text = xml_generator.build_xml_string(seq, port_id=port)
        parsed = xml_generator.parse_xml_string(xml_text) or {}
        parsed_seq = str(parsed.get("sequence_name", "")).strip().upper()
        if parsed_seq:
            # XML이 알려주는 정식 sequence를 최우선으로 채택
            seq_for_mapping = parsed_seq
    except Exception as e:
        if verbose:
            print(f"[ANIM HOOK] XML 생성/역파싱 실패: seq={seq}, err={e}", flush=True)
        return

    # rules/map 입력을 XML 역파싱 기준으로 완전히 표준화
    # from/to/port 텍스트는 원본 접두사(BP/EP/OHT) 힌트를 살려 재구성한다.
    mapping_payload = dict(payload or {})
    parsed_from = str(parsed.get("from_port_id", "") or "")
    parsed_to = str(parsed.get("to_port_id", "") or "")
    parsed_port = str(parsed.get("port_id", "") or "")
    mapping_payload["seq"] = seq_for_mapping
    mapping_payload["from_port_id"] = _normalize_port_text_from_xml(parsed_from, from_port_txt)
    mapping_payload["to_port_id"] = _normalize_port_text_from_xml(parsed_to, to_port_txt)
    mapping_payload["port_id"] = _normalize_port_text_from_xml(parsed_port, port_txt)
    mapping_payload["_xml_sequence_name"] = seq_for_mapping
    mapping_payload["_xml_text"] = xml_text
    mapping_payload["_xml_parsed"] = parsed

    # 역파싱된 seq를 기준으로 기존 rules/map(JSON 파일)은 그대로 사용
    # 즉, 규칙 파일은 수정하지 않고도 XML 표준화 파이프라인 위에서 동작한다.
    mapped_json, mapped_meta, matched_rule, matched_source = _resolve_event_animation_entry(seq_for_mapping, mapping_payload)
    if mapped_json:
        _append_anim_history_log(
            ext,
            f"[ANIM MAP] source={matched_source or '-'} rule={matched_rule or '-'} event={seq_for_mapping} file={Path(str(mapped_json)).name}",
        )
        _execute_mapped_sequence_stub(ext, seq_for_mapping, mapping_payload, mapped_json, mapped_meta, matched_rule, verbose)
    elif verbose:
        print(
            f"[ANIM MAP] 이벤트={seq_for_mapping} 매핑 없음 "
            f"(config/event_animation_rules.json / event_animation_map.json 확인)",
            flush=True,
        )
    if not mapped_json:
        hint_name = f"{seq_for_mapping.lower()}.json".replace("eapeis_port_", "")
        _append_anim_history_log(
            ext,
            f"[ANIM] 매핑없음 | event={seq_for_mapping} | 필요한 예시파일={hint_name}",
        )
    # 이후 출력도 XML 표준화된 값 기준으로 일관성 유지
    from_port_txt = str(mapping_payload.get("from_port_id", ""))
    to_port_txt = str(mapping_payload.get("to_port_id", ""))
    port_txt = str(mapping_payload.get("port_id", ""))
    from_kind = _port_kind(from_port_txt)
    to_kind = _port_kind(to_port_txt)
    port_kind = _port_kind(port_txt)

    action_desc = parsed.get("action_desc", "")
    if action_desc and verbose:
        if port_txt:
            action_desc += f" | 대상포트={port_txt}({port_kind})"
        if from_port_txt or to_port_txt:
            action_desc += f" | 이동경로={from_port_txt}({from_kind})->{to_port_txt}({to_kind})"
        print(f"[ANIM HOOK ACTION] {action_desc}", flush=True)

    # 추후 실제 애니메이션 분기 지점
    if seq_for_mapping == xml_generator.SEQ_READYTOLOAD:
        if verbose:
            print(f"[ANIM PLAN] READY_TO_LOAD 대기 상태 애니메이션 | port={port_txt}({port_kind})", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_ARRIVED:
        if verbose:
            print(f"[ANIM PLAN] ARRIVED 안착 애니메이션 | port={port_txt}({port_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_MOVE_TRANSFERING:
        if verbose:
            print(f"[ANIM PLAN] MOVE_TRANSFERING 이송 애니메이션 | from={from_port_txt}({from_kind}) to={to_port_txt}({to_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_MOVE:
        if verbose:
            print(f"[ANIM PLAN] MOVE 이동 애니메이션 | from={from_port_txt}({from_kind}) to={to_port_txt}({to_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_READYTOUNLOAD:
        if verbose:
            print(f"[ANIM PLAN] READY_TO_UNLOAD 회수 준비 애니메이션 | port={port_txt}({port_kind}) lot={lot_id}", flush=True)
    elif seq_for_mapping == xml_generator.SEQ_REMOVED:
        if verbose:
            print(f"[ANIM PLAN] REMOVED 회수 완료 애니메이션 | port={port_txt}({port_kind}) lot={lot_id}", flush=True)
    else:
        if verbose:
            print(f"[ANIM PLAN] 미분류 이벤트 | seq={seq_for_mapping} payload={payload}", flush=True)


def _is_progress_only_mode(ext: Any) -> bool:
    try:
        return ext._sim_log_view_combo.model.get_item_value_model().as_int == int(SimLogPanelMode.PROGRESS_ONLY)
    except Exception:
        return False


def _export_sim_logs_to_xlsx(ext: Any) -> None:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        _append_sim_log(ext, f"[SIM EXPORT] openpyxl import 실패: {e}")
        return

    out_dir = _extension_root_dir() / "data" / "sim_logs"
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"sim_logs_{ts}.xlsx"

    def _rows(text: str) -> List[str]:
        lines = [ln for ln in (text or "").splitlines() if ln.strip()]
        # 실행 순서(오래된 항목 -> 최신 항목) 그대로 저장
        return lines

    progress_rows = _rows(getattr(ext, "_sim_progress_label", None).text if getattr(ext, "_sim_progress_label", None) else "")
    history_rows = _rows(getattr(ext, "_sim_history_label", None).text if getattr(ext, "_sim_history_label", None) else "")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "진행현황"
    ws2 = wb.create_sheet("이력로그")

    for idx, row in enumerate(progress_rows, start=1):
        ws1.cell(row=idx, column=1, value=row)
    for idx, row in enumerate(history_rows, start=1):
        ws2.cell(row=idx, column=1, value=row)

    wb.save(str(out_path))
    _append_sim_log(ext, f"[SIM EXPORT] 저장 완료: {out_path}")


def _detach_sim_update(ext: Any) -> None:
    sub = getattr(ext, "_sim_update_sub", None)
    if sub is not None:
        try:
            sub.unsubscribe()
        except Exception:
            pass
        ext._sim_update_sub = None

    stop_evt = getattr(ext, "_sim_thread_stop", None)
    th = getattr(ext, "_sim_thread", None)
    if stop_evt is not None:
        try:
            stop_evt.set()
        except Exception:
            pass
    if th is not None:
        try:
            th.join(timeout=1.0)
        except Exception:
            pass
    ext._sim_thread = None
    ext._sim_thread_stop = None
    ui_sub = getattr(ext, "_sim_log_ui_sub", None)
    if ui_sub is not None:
        try:
            ui_sub.unsubscribe()
        except Exception:
            pass
        ext._sim_log_ui_sub = None


def on_sim_start_clicked(ext: Any) -> None:
    try:
        ep_count_idx = ext._sim_ep_count_combo.model.get_item_value_model().as_int
    except Exception:
        ep_count_idx = 0
    ep_count = 2 if ep_count_idx == 0 else 3

    initial_full_ports: List[str] = []
    if ext._sim_init_bp1_model.get_value_as_bool():
        initial_full_ports.append("BP1")
    if ext._sim_init_bp2_model.get_value_as_bool():
        initial_full_ports.append("BP2")
    if ext._sim_init_bp3_model.get_value_as_bool():
        initial_full_ports.append("BP3")
    if ext._sim_init_bp4_model.get_value_as_bool():
        initial_full_ports.append("BP4")
    if ext._sim_init_ep1_model.get_value_as_bool():
        initial_full_ports.append("EP1")
    if ext._sim_init_ep2_model.get_value_as_bool():
        initial_full_ports.append("EP2")
    if ep_count >= 3 and ext._sim_init_ep3_model.get_value_as_bool():
        initial_full_ports.append("EP3")

    on_sim_stop_clicked(ext)
    lot_count = max(1, ext._sim_lot_count_model.get_value_as_int())
    spawn_imin = max(0.1, ext._sim_lot_spawn_min_model.get_value_as_float())
    spawn_imax = max(0.1, ext._sim_lot_spawn_max_model.get_value_as_float())
    if spawn_imin > spawn_imax:
        spawn_imin, spawn_imax = spawn_imax, spawn_imin
    pue_min = max(0.1, ext._sim_pickup_evt_min_model.get_value_as_float())
    pue_max = max(0.1, ext._sim_pickup_evt_max_model.get_value_as_float())
    if pue_min > pue_max:
        pue_min, pue_max = pue_max, pue_min
    timing = SimulationTimingConfig(
        oht_to_bp1_min=max(0.1, ext._sim_oht_bp1_min_model.get_value_as_float()),
        oht_to_bp1_max=max(0.1, ext._sim_oht_bp1_max_model.get_value_as_float()),
        bp1_to_bp_min=max(0.1, ext._sim_bp1_bp_min_model.get_value_as_float()),
        bp1_to_bp_max=max(0.1, ext._sim_bp1_bp_max_model.get_value_as_float()),
        bp_to_ep_min=max(0.1, ext._sim_bp_ep_min_model.get_value_as_float()),
        bp_to_ep_max=max(0.1, ext._sim_bp_ep_max_model.get_value_as_float()),
        ep_to_oht_min=max(0.1, ext._sim_ep_oht_min_model.get_value_as_float()),
        ep_to_oht_max=max(0.1, ext._sim_ep_oht_max_model.get_value_as_float()),
        lot_spawn_interval_min=spawn_imin,
        lot_spawn_interval_max=spawn_imax,
        pickup_event_interval_min=pue_min,
        pickup_event_interval_max=pue_max,
    )
    log_interval = max(0.0, ext._sim_log_interval_model.get_value_as_float())
    log_cfg = SimulationLogConfig(
        progress_interval_sec=log_interval,
        input_status_interval_sec=log_interval,
    )
    init_cfg = SimulationInitConfig(
        ep_count=ep_count,
        initial_full_ports=initial_full_ports,
        max_oht_lots=lot_count,
    )
    lots: List[Lot] = []

    ext._sim_history_text.set_value("[SIM] 초기화")
    ext._sim_progress_text.set_value("[진행현황] 초기화 (시뮬레이션 시작 대기)")
    ext._sim_port_state_text.set_value("[포트상태] 초기화 (이벤트 대기)")
    if getattr(ext, "_sim_history_label", None) is not None:
        ext._sim_history_label.text = "[SIM] 초기화"
    if getattr(ext, "_sim_progress_label", None) is not None:
        ext._sim_progress_label.text = "[진행현황] 초기화 (시뮬레이션 시작 대기)"
    if getattr(ext, "_sim_port_state_label", None) is not None:
        ext._sim_port_state_label.text = "[포트상태] 초기화 (이벤트 대기)"
    if getattr(ext, "_sim_port_state_header_label", None) is not None:
        ext._sim_port_state_header_label.text = "[포트상태] 초기화 (이벤트 대기)"
    cells = getattr(ext, "_sim_port_cells", {}) or {}
    for port in ("BP2", "BP3", "BP4", "BP1", "EP1", "EP2"):
        if port in cells:
            cells[port].text = f"{port}:-"
    if getattr(ext, "_sim_port_ep3_cell", None) is not None:
        ext._sim_port_ep3_cell.text = "EP3:-"
    ext._sim_progress_rows = {}
    ext._sim_progress_history = []
    ext._sim_progress_start_times = {}
    ext._sim_log_queue = queue.SimpleQueue()
    _enqueue_sim_log(ext, "[SIM UI] 실시간 로그 큐 초기화")
    ext._sim_anim_active = {}
    ext._sim_anim_pending = []

    def _on_gate(payload: Dict[str, str]) -> float:
        # 요구사항: 공정시간보다 애니(JSON) 시간이 길면 다음 공정은 애니 종료까지 대기.
        # simulation_engine은 이 반환값(초)을 받아서 각 공정 timeout을 max(공정, 애니)로 확장한다.
        anim_est_sec = _estimate_anim_duration_for_gate_payload(ext, payload or {})
        # 공정확인 체크 시에는 "확인 클릭 전에는 애니/공정 시작 금지"가 목표이므로,
        # simulation_engine의 _request_gate() 시점에 UI 확인창을 띄우고 동기 블로킹한다.
        try:
            confirm_each = bool(
                getattr(ext, "_sim_confirm_each_step_model", None) is not None
                and ext._sim_confirm_each_step_model.get_value_as_bool()
            )
        except Exception:
            confirm_each = False
        if not confirm_each:
            return float(anim_est_sec)

        # 공정확인 중에는 sim tick thread도 멈춰야 "확인 전까지 완전 정지"가 된다.
        try:
            gp = getattr(ext, "_sim_gate_pause_event", None)
            if gp is not None:
                gp.set()
        except Exception:
            pass

        seq_raw = str(payload.get("seq", ""))
        seq = SIM_SEQ_ALIAS.get(seq_raw, seq_raw)
        lot = str(payload.get("lot_id", ""))
        est = str(payload.get("est_sec", ""))
        fr = str(payload.get("from_port_id", ""))
        to = str(payload.get("to_port_id", ""))
        port = str(payload.get("port_id", ""))
        try:
            # 게이트 다이얼로그의 XML도 실제 애니 매핑 파이프라인과 같은 규칙으로 생성한다.
            # (FROM/TO 시퀀스 vs PORT_ID 시퀀스 분기)
            if seq in xml_generator.FROM_TO_SEQS:
                fnum = _parse_port_num(fr, 1)
                tnum = _parse_port_num(to, 1)
                xml = xml_generator.build_xml_string(seq, from_port_id=fnum, to_port_id=tnum)
            elif seq in xml_generator.PORT_ID_ONLY_SEQS:
                pnum = _parse_port_num(port, 1)
                xml = xml_generator.build_xml_string(seq, port_id=pnum)
            else:
                xml = f"(XML 미생성: 비-XML 공정 seq={seq_raw}->{seq})"
        except Exception:
            xml = f"(XML 생성 실패: seq={seq})"

        xml_sequence_name = ""
        if isinstance(xml, str) and xml.strip().startswith("<"):
            try:
                _pd = xml_generator.parse_xml_string(xml) or {}
                xml_sequence_name = str(_pd.get("sequence_name", "") or "").strip().upper()
            except Exception:
                pass

        # Alert에서 "실행 대상 JSON 파일"과 존재 여부를 함께 안내한다.
        map_line = "JSON 매핑: 없음"
        try:
            mapping_payload = dict(payload or {})
            seq_for_mapping = seq
            if isinstance(xml, str) and xml.strip().startswith("<"):
                parsed = xml_generator.parse_xml_string(xml) or {}
                parsed_seq = str(parsed.get("sequence_name", "") or "").strip().upper()
                if parsed_seq:
                    seq_for_mapping = parsed_seq
                mapping_payload["seq"] = seq_for_mapping
                mapping_payload["from_port_id"] = _normalize_port_text_from_xml(str(parsed.get("from_port_id", "") or ""), fr)
                mapping_payload["to_port_id"] = _normalize_port_text_from_xml(str(parsed.get("to_port_id", "") or ""), to)
                mapping_payload["port_id"] = _normalize_port_text_from_xml(str(parsed.get("port_id", "") or ""), port)
            else:
                mapping_payload["seq"] = seq_for_mapping

            mapped_json, _meta, rule_name, source_name = _resolve_event_animation_entry(seq_for_mapping, mapping_payload)
            if mapped_json:
                jp = _normalize_json_path(mapped_json)
                exists_txt = "존재" if jp.exists() else "없음"
                map_line = (
                    f"JSON 매핑: source={source_name or '-'} rule={rule_name or '-'} "
                    f"file={jp.name} ({exists_txt})"
                )
            else:
                map_line = f"JSON 매핑: 없음 (event={seq_for_mapping})"
        except Exception as e:
            map_line = f"JSON 매핑 확인 실패: {e}"
        done_evt = threading.Event()
        message = (
            f"공정: {payload.get('title','-')}\n"
            f"이벤트 sequence_name: 시뮬 seq={seq_raw or '-'}, 규격/별칭={seq or '-'}"
            + (f", XML SEQUENCE_NAME={xml_sequence_name}" if xml_sequence_name else "")
            + "\n"
            f"lot={lot} from={fr} to={to} port={port}\n"
            f"예상시간={est}s\n"
            f"애니예상={anim_est_sec:.2f}s (JSON 기준)\n\n"
            f"{map_line}\n\n"
            f"XML:\n{xml}"
        )
        _enqueue_gate_request(
            ext,
            {
                "title": payload.get("title", "공정 확인"),
                "message": message,
                "_done_event": done_evt,
                "gate_seq_raw": seq_raw,
                "gate_seq_canonical": seq,
                "gate_xml_sequence_name": xml_sequence_name,
            },
        )
        # 시뮬레이션 스레드는 사용자 확인 전까지 여기서 동기 대기한다.
        done_evt.wait()
        return float(anim_est_sec)

    engine = TBSSimulationEngine(
        lots=lots,
        timing=timing,
        log_config=log_cfg,
        init_config=init_cfg,
        # 시뮬레이션 스레드에서 발생하는 로그/이벤트는 큐에 넣고 UI 스레드에서만 렌더링한다.
        # (Omni UI 스레드 제약 회피)
        on_log=lambda line: post_sim_history_line(ext, line),
        on_event=lambda payload: post_sim_anim_event(ext, payload),
        on_progress=lambda payload: post_sim_progress_update(ext, payload),
        on_gate=_on_gate,
        print_to_console=(not _is_progress_only_mode(ext)),
    )
    ext._sim_engine = engine
    if not engine.start():
        _append_sim_log(ext, "[SIM] 시작 실패")
        return

    tick_state = {"count": 0}
    stop_evt = threading.Event()
    ext._sim_thread_stop = stop_evt
    speed_value = max(0.1, ext._sim_speed_model.get_value_as_float())
    _append_sim_log(ext, f"[SIM] tick thread 준비 (speed={speed_value:.2f}x)")
    try:
        ext._sim_log_ui_sub = app.get_app().get_update_event_stream().create_subscription_to_pop(
            lambda e: _drain_sim_log_queue(ext),
            name="morph.tbs_control_1:sim_log_ui_drain",
        )
    except Exception as e:
        _append_sim_log(ext, f"[SIM UI] 로그 큐 드레인 구독 실패: {e}")

    def _tick_loop():
        try:
            print("[SIM] tick thread 시작", flush=True)
            last = time.perf_counter()
            while not stop_evt.is_set():
                # 애니메이션이 재생 중이면 sim tick을 일시정지
                pause_evt = getattr(ext, "_sim_tick_pause_event", None)
                gate_pause_evt = getattr(ext, "_sim_gate_pause_event", None)
                try:
                    confirm_each = bool(getattr(ext, "_sim_confirm_each_step_model", None) is not None and ext._sim_confirm_each_step_model.get_value_as_bool())
                except Exception:
                    confirm_each = False
                if not confirm_each and gate_pause_evt is not None and gate_pause_evt.is_set():
                    try:
                        gate_pause_evt.clear()
                    except Exception:
                        pass
                if confirm_each and gate_pause_evt is not None and gate_pause_evt.is_set():
                    time.sleep(0.02)
                    continue
                if pause_evt is not None and pause_evt.is_set():
                    # 원칙: JSON 애니메이션이 실제로 진행 중이면(sim 모듈의 활성 상태가 있으면) 절대 tick 재개하지 않는다.
                    try:
                        anim_running = bool(
                            translate_animation.is_translate_animation_running()
                            or rotate_animation.is_rotate_animation_running()
                            or curve_animation.is_curve_animation_running()
                            or (getattr(ext, "_sim_runner", None) is not None and getattr(ext._sim_runner, "is_running", lambda: False)())
                        )
                    except Exception:
                        anim_running = True
                    if anim_running:
                        time.sleep(0.02)
                        continue

                    # fail-safe: 추정 시간이 남아있으면 최소한 그동안은 pause 유지
                    until_wall = getattr(ext, "_sim_tick_pause_until_wall", None)
                    if isinstance(until_wall, (float, int)) and time.monotonic() < float(until_wall):
                        time.sleep(0.02)
                        continue

                    time.sleep(0.02)
                    continue
                sim = getattr(ext, "_sim_engine", None)
                if sim is None:
                    break
                now = time.perf_counter()
                dt = now - last
                last = now
                dt = max(0.001, min(dt, 0.1))
                sim.tick(dt * speed_value)
                tick_state["count"] += 1
                if tick_state["count"] == 1:
                    print("[SIM] tick 동작 확인 (first tick)", flush=True)
                if sim.is_done:
                    print("[SIM] 종료 감지", flush=True)
                    _enqueue_control_action(ext, SimUiControlAction.EXPORT_XLSX.value)
                    break
                time.sleep(0.02)
        except Exception as err:
            # 원인 파악을 위해 traceback까지 출력한다.
            try:
                import traceback

                print(f"[SIM] tick thread 예외: {err}", flush=True)
                print(traceback.format_exc(), flush=True)
            except Exception:
                print(f"[SIM] tick thread 예외: {err}", flush=True)

    th = threading.Thread(target=_tick_loop, name="morph.tbs_control_1.sim_tick", daemon=True)
    ext._sim_thread = th
    th.start()


def on_sim_stop_clicked(ext: Any) -> None:
    sim = getattr(ext, "_sim_engine", None)
    if sim is not None:
        try:
            sim.stop()
        except Exception:
            pass
    _detach_sim_update(ext)
    runner = getattr(ext, "_sim_runner", None)
    if runner is not None:
        try:
            runner.stop()
        except Exception:
            pass
    # pause 상태 해제
    pe = getattr(ext, "_sim_tick_pause_event", None)
    if pe is not None:
        try:
            pe.clear()
        except Exception:
            pass
    ge = getattr(ext, "_sim_gate_pause_event", None)
    if ge is not None:
        try:
            ge.clear()
        except Exception:
            pass
    # 공정확인 창이 열려있으면 강제 종료(리셋/중지 시 다음 실행에 pause가 남지 않게)
    try:
        w = getattr(ext, "_sim_gate_dialog", None)
        if w is not None:
            try:
                w.visible = False
            except Exception:
                pass
            try:
                w.destroy()
            except Exception:
                pass
        ext._sim_gate_dialog = None
    except Exception:
        pass
    try:
        ext._sim_tick_pause_until_wall = None
    except Exception:
        pass
    try:
        ext._sim_anim_pending = []
    except Exception:
        pass


def on_sim_reset_clicked(ext: Any) -> None:
    on_sim_stop_clicked(ext)
    try:
        clear_port_lot_authoring_cache()
    except Exception:
        pass
    ext._sim_engine = None
    if getattr(ext, "_sim_history_text", None):
        ext._sim_history_text.set_value("[SIM] 리셋 완료")
    if getattr(ext, "_sim_history_label", None) is not None:
        ext._sim_history_label.text = "[SIM] 리셋 완료"
    if getattr(ext, "_sim_progress_text", None):
        ext._sim_progress_text.set_value("[진행현황] 없음")
    if getattr(ext, "_sim_progress_label", None) is not None:
        ext._sim_progress_label.text = "[진행현황] 없음"
    if getattr(ext, "_sim_port_state_text", None):
        ext._sim_port_state_text.set_value("[포트상태] 없음")
    if getattr(ext, "_sim_port_state_label", None) is not None:
        ext._sim_port_state_label.text = "[포트상태] 없음"
    if getattr(ext, "_sim_port_state_header_label", None) is not None:
        ext._sim_port_state_header_label.text = "[포트상태] 없음"
    cells = getattr(ext, "_sim_port_cells", {}) or {}
    for port in ("BP2", "BP3", "BP4", "BP1", "EP1", "EP2"):
        if port in cells:
            cells[port].text = f"{port}:-"
        # 포트 박스 배경색(점유 색상)도 EMPTY로 초기화
        try:
            _set_port_box_style(ext, port, "-")
        except Exception:
            pass
    if getattr(ext, "_sim_port_ep3_cell", None) is not None:
        ext._sim_port_ep3_cell.text = "EP3:-"
    try:
        _set_port_box_style(ext, "EP3", "-")
    except Exception:
        pass
    ext._sim_progress_rows = {}
    ext._sim_progress_history = []
    ext._sim_progress_start_times = {}
    # 최근 요약/대기 토큰 초기화
    try:
        ext._sim_recent_story_blocks = []
    except Exception:
        pass
    # (요청으로 제거) 점 표시 기능 비활성화


def on_sim_log_view_changed(ext: Any) -> None:
    try:
        idx = ext._sim_log_view_combo.model.get_item_value_model().as_int
    except Exception:
        idx = int(SimLogPanelMode.ALL)
    # 구버전 콤보(4항목) 저장값 호환: 인덱스 3 이상은 둘다로 취급
    if idx > int(SimLogPanelMode.HISTORY_ONLY):
        idx = int(SimLogPanelMode.ALL)
    try:
        mode = SimLogPanelMode(int(idx))
    except Exception:
        mode = SimLogPanelMode.ALL
    if getattr(ext, "_sim_progress_frame", None) is not None:
        ext._sim_progress_frame.visible = mode in (SimLogPanelMode.ALL, SimLogPanelMode.PROGRESS_ONLY)
    if getattr(ext, "_sim_history_frame", None) is not None:
        ext._sim_history_frame.visible = mode in (SimLogPanelMode.ALL, SimLogPanelMode.HISTORY_ONLY)
    sim = getattr(ext, "_sim_engine", None)
    if sim is not None and hasattr(sim, "set_console_logging_enabled"):
        # 진행현황 전용 모드에서는 콘솔/이력 로그 최소화
        sim.set_console_logging_enabled(mode != SimLogPanelMode.PROGRESS_ONLY)


def on_copy_sim_progress(ext: Any) -> None:
    text = ""
    if getattr(ext, "_sim_progress_label", None) is not None:
        text = ext._sim_progress_label.text or ""
    if not text.strip() and getattr(ext, "_sim_progress_text", None):
        text = ext._sim_progress_text.as_string or ""
    if not text.strip():
        _append_sim_log(ext, "[SIM UI] 복사할 진행현황이 없습니다.")
        return
    try:
        import omni.kit.clipboard as cb  # type: ignore
        if hasattr(cb, "copy"):
            cb.copy(text)
        elif hasattr(cb, "set_text"):
            cb.set_text(text)
        else:
            raise RuntimeError("clipboard api not found")
        _append_sim_log(ext, "[SIM UI] 진행현황 복사 완료")
    except Exception:
        print("[SIM UI] 클립보드 미지원: 진행현황을 콘솔에 출력합니다.", flush=True)
        print(text, flush=True)
        _append_sim_log(ext, "[SIM UI] 클립보드 미지원으로 콘솔 출력")


def on_sim_ep_count_changed(ext: Any) -> None:
    try:
        idx = ext._sim_ep_count_combo.model.get_item_value_model().as_int
    except Exception:
        idx = 0
    is_ep3 = idx == 1
    if getattr(ext, "_sim_init_ep3_row", None) is not None:
        ext._sim_init_ep3_row.visible = is_ep3
    if not is_ep3 and getattr(ext, "_sim_init_ep3_model", None) is not None:
        ext._sim_init_ep3_model.set_value(False)
    _sync_ep3_port_cell_visibility(ext)


def receive_signal_data(ext: Any, data: str, format: str = "json") -> bool:
    parsed = parse_signal(data, format)
    if not parsed:
        return False
    run_generator_from_parsed(ext, parsed)
    return True


def run_generator_from_parsed(ext: Any, parsed: dict) -> None:
    stage = get_stage()
    if not stage:
        return
    objects = parsed.get("objects") or []
    segments = parsed.get("segments") or []
    if not objects or not segments:
        return
    for name in objects:
        if not isinstance(name, str):
            continue
        paths = find_all_prim_paths_by_name(stage, name)
        for path in paths:
            if not path:
                continue
            stop_prim_translate_animation(path)
            stop_prim_curve_animation(path)
            run_prim_translate_animation(path, segments, loop=False)


def on_refresh_prim_list(ext: Any) -> None:
    stage = get_stage()
    if not stage:
        if getattr(ext, "_load_status_label", None):
            ext._load_status_label.text = "스테이지가 없습니다. USD를 먼저 로드하세요."
        return
    ext._tracked_paths = collect_prim_paths_safe(stage)
    refresh_object_list(ext)


def refresh_object_list(ext: Any) -> None:
    if ext._object_list_frame is None:
        return
    ext._object_list_frame.clear()
    stage = get_stage()
    if not stage:
        with ext._object_list_frame:
            ui.Label("USD를 먼저 로드하세요.")
        return

    def _valid_path(p: str) -> bool:
        try:
            return stage.GetPrimAtPath(p).IsValid()
        except (UnicodeDecodeError, UnicodeEncodeError):
            return False

    valid_paths = [p for p in ext._tracked_paths if _valid_path(p)]
    total = len(valid_paths)
    priority_prefix = (ext._priority_prefix_model.get_value_as_string().strip() or "")

    if priority_prefix:
        priority_paths: List[str] = []
        rest_paths: List[str] = []
        for p in valid_paths:
            try:
                prim = stage.GetPrimAtPath(p)
                if not prim or not prim.IsValid():
                    rest_paths.append(p)
                    continue
                name = safe_str(prim.GetName())
                if name.startswith(priority_prefix):
                    priority_paths.append(p)
                else:
                    rest_paths.append(p)
            except Exception:
                rest_paths.append(p)
        need = max(0, MAX_PRIMS_DISPLAY - len(priority_paths))
        display_paths = priority_paths[:MAX_PRIMS_DISPLAY] + rest_paths[:need]
    else:
        display_paths = valid_paths[:MAX_PRIMS_DISPLAY]

    with ext._object_list_frame:
        if total > MAX_PRIMS_DISPLAY:
            ui.Label(f"총 {total}개 prim 중 {len(display_paths)}개만 표시됩니다.", height=0)
            ui.Spacer(height=4)
        if priority_prefix:
            n_priority = min(len(priority_paths), MAX_PRIMS_DISPLAY)
            n_rest = len(display_paths) - n_priority
            ui.Label(f"접두사 '{priority_prefix}' 우선: {n_priority}개, 나머지 순서대로 {n_rest}개", height=0)
            ui.Spacer(height=4)
        for idx, prim_path in enumerate(display_paths):
            build_object_panel(ext, ext._object_list_frame, prim_path, idx + 1)


def build_object_panel(ext: Any, parent: ui.VStack, prim_path: str, index: int) -> None:
    try:
        stage = get_stage()
        prim = stage.GetPrimAtPath(prim_path) if stage else None
        if not prim or not prim.IsValid():
            return
        title = get_prim_display_name(prim, index)
        local = get_prim_local_translate(prim)
        pos_models = [
            ui.SimpleFloatModel(local[0]),
            ui.SimpleFloatModel(local[1]),
            ui.SimpleFloatModel(local[2]),
        ]

        def update_prim_position():
            s = get_stage()
            p = s.GetPrimAtPath(prim_path) if s else None
            if p and p.IsValid():
                set_prim_translate_only(p, Gf.Vec3f(
                    pos_models[0].get_value_as_float(),
                    pos_models[1].get_value_as_float(),
                    pos_models[2].get_value_as_float(),
                ))

        with parent:
            with ui.CollapsableFrame(title, collapsed=False):
                with ui.VStack(spacing=6):
                    ui.Label("Position (X, Y, Z)", height=0)
                    with ui.HStack():
                        for i, label in enumerate(["X", "Y", "Z"]):
                            ui.Label(label, width=24)
                            ui.FloatField(model=pos_models[i])
                    for m in pos_models:
                        m.add_value_changed_fn(lambda _: update_prim_position())
                    ui.Spacer(height=4)
                    ui.Button("3D 정보 보기", height=24, clicked_fn=lambda p=prim_path: show_prim_info_in_viewport(ext, p))
                    ui.Spacer(height=4)
                    with ui.HStack(spacing=8):
                        ui.Button("button_0", width=0, clicked_fn=lambda p=prim_path: on_button_0(ext, p))
                        ui.Button("button_1", width=0, clicked_fn=lambda p=prim_path: on_button_1(ext, p))
                        ui.Button("button_2", width=0, clicked_fn=lambda p=prim_path: on_button_2(ext, p))
    except (UnicodeDecodeError, UnicodeEncodeError):
        return


def on_button_0(ext: Any, prim_path: str) -> None:
    stop_prim_translate_animation(prim_path)
    stop_prim_curve_animation(prim_path)
    stop_prim_rotate_animation(prim_path)
    run_prim_translate_animation(
        prim_path,
        [
            {"duration": 1.0, "delta": (100.0, 0.0, 0.0)},
            {"duration": 1.0, "delta": (0.0, 0.0, 100.0)},
        ],
        loop=False,
    )


def on_button_1(ext: Any, prim_path: str) -> None:
    stop_prim_translate_animation(prim_path)
    stop_prim_curve_animation(prim_path)
    stop_prim_rotate_animation(prim_path)
    stage = get_stage()
    prim = stage.GetPrimAtPath(prim_path) if stage else None
    if not prim or not prim.IsValid():
        return
    start = get_prim_local_translate(prim)
    start_t = (start[0], start[1], start[2])
    end_t = (start[0] + 100.0, start[1], start[2])
    path_points = make_parabolic_path(start=start_t, end=end_t, arc_height=30.0, num_points=24)
    run_prim_curve_animation(prim_path, path_points, duration_sec=1.0, loop=False)


def on_button_2(ext: Any, prim_path: str) -> None:
    stop_prim_translate_animation(prim_path)
    stop_prim_curve_animation(prim_path)
    stop_prim_rotate_animation(prim_path)
    run_prim_rotate_animation(
        prim_path,
        [{"duration": 3.0, "delta": (0.0, 90.0, 0.0)}],
        loop=False,
    )
